"""
自社在庫（Raw materials daily check）とTTAF在庫（CSA ReportのPIVOT SOH TTAFシート）の
週次実績を抽出する。

使い方:
    python3 scripts/extract_self_ttaf_stock.py self <daily_check.xlsxのパス> [YYYY-MM-DD]
    python3 scripts/extract_self_ttaf_stock.py ttaf <CSA_Report.xlsxのパス> [YYYY-MM-DD]

日付を省略した場合、ファイル名から DD.MM.YYYY 形式の日付を自動抽出します
（見つからない場合はエラーになるので、その場合は明示的に指定してください）。

抽出結果は data/masters/self_stock_sample.csv / ttaf_stock_sample.csv に**追記**されます
（同じ RM_Code+Date の組み合わせがあれば上書き）。build_soh.py を再実行するとSOH_Master.xlsx
に反映されます。通常の月次運用ではVBAマクロ（RefreshSelfStock/RefreshTTAFStock）を使う想定で、
このPythonスクリプトはPython環境がある場合の代替手段・検証用です。
"""
import openpyxl, csv, sys, os, re, datetime

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MASTERS = os.path.join(REPO_ROOT, "data", "masters")


def parse_date_from_filename(path):
    m = re.search(r"(\d{2})\.(\d{2})\.(\d{4})", os.path.basename(path))
    if not m:
        raise ValueError(f"ファイル名から日付を自動抽出できませんでした: {path}\n"
                          f"第3引数にYYYY-MM-DD形式で日付を指定してください。")
    d, mth, y = m.groups()
    return datetime.date(int(y), int(mth), int(d))


def extract_self(src_path, dt):
    wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
    ws = wb["Stock"]
    rows = []
    for r in ws.iter_rows(min_row=9, max_row=90, max_col=10, values_only=True):
        code, total = r[2], r[9]
        if isinstance(code, str) and code.startswith("CHEM") and isinstance(total, (int, float)):
            rows.append([code, dt.isoformat(), total])
    wb.close()
    return rows


def extract_ttaf(src_path, dt):
    wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
    ws = wb["PIVOT SOH TTAF"]
    rows = []
    for row in ws.iter_rows(min_row=5, max_col=4, values_only=True):
        if row[0] == "Grand Total" or row[0] is None:
            continue
        code, aged, fresh, total = row
        if isinstance(code, str) and code.startswith("CHEM") and isinstance(total, (int, float)):
            rows.append([code, dt.isoformat(), total])
    wb.close()
    return rows


def merge_into(csv_path, header, new_rows):
    existing = {}
    if os.path.exists(csv_path):
        with open(csv_path, newline="", encoding="utf-8") as f:
            for row in csv.reader(f):
                if row and row[0] != header[0]:
                    existing[(row[0], row[1])] = row
    for row in new_rows:
        existing[(row[0], row[1])] = row
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(header)
        for row in sorted(existing.values(), key=lambda r: (r[0], r[1])):
            w.writerow(row)
    print(f"{csv_path}: {len(existing)} rows total")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    kind, src = sys.argv[1], sys.argv[2]
    dt = datetime.date.fromisoformat(sys.argv[3]) if len(sys.argv) > 3 else parse_date_from_filename(src)

    if kind == "self":
        rows = extract_self(src, dt)
        merge_into(os.path.join(MASTERS, "self_stock_sample.csv"), ["RM_Code", "Date", "Self_Qty"], rows)
    elif kind == "ttaf":
        rows = extract_ttaf(src, dt)
        merge_into(os.path.join(MASTERS, "ttaf_stock_sample.csv"), ["RM_Code", "Date", "TTAF_Qty"], rows)
    else:
        print("第1引数は 'self' か 'ttaf' を指定してください")
        sys.exit(1)

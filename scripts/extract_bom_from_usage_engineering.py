"""
「Usage from Production Engineering」（実ファイル名は "Powder & Slurry & Pgm Plan" 系列とは別の、
原単位専用ファイル）の Slurry / Powder シートから原単位(1バッチあたり使用量kg)を抽出する。

シート構造: A列=中間体名、2行目(ヘッダー)=材料名、交わるセル=1バッチ使用量(kg)。

使い方:
    python3 scripts/extract_bom_from_usage_engineering.py <入力xlsxパス> [出力ディレクトリ]

出力先は既定で data/masters/bom.csv （このスクリプト単独の出力を主として使う場合は上書き、
Powder & Slurry & Pgm Plan 側の bom_from_plan.csv と合わせて使う場合は merge_bom_sources.py を
別途実行してください）。
"""
import openpyxl, csv, sys, os


def norm(s):
    return "".join(ch for ch in s.upper() if ch.isalnum())


def extract(src_path, out_dir):
    wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)
    bom = {}

    def harvest(sheet_name):
        if sheet_name not in wb.sheetnames:
            print(f"  [skip] sheet {sheet_name!r} not found")
            return
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[1]
        for r in rows[3:]:
            inter = r[0]
            if not inter or not isinstance(inter, str):
                continue
            inter = inter.strip()
            for ci in range(1, len(r)):
                val = r[ci]
                if val is None or ci >= len(header) or not header[ci]:
                    continue
                mat_name = str(header[ci]).strip()
                if isinstance(val, (int, float)) and val != 0:
                    bom[(inter, mat_name)] = float(val)

    harvest("Slurry")
    harvest("Powder")
    wb.close()
    print("bom pairs (by material name):", len(bom))

    rm_master_path = os.path.join(out_dir, "rm_master.csv")
    desc_index = {}
    if os.path.exists(rm_master_path):
        with open(rm_master_path) as f:
            for r in csv.DictReader(f):
                desc_index[norm(r["Description"])] = r["RM_Code"]

    resolved = []
    unresolved = set()
    for (inter, mat_name), qty in bom.items():
        key = norm(mat_name)
        rm_code = desc_index.get(key)
        if rm_code is None:
            for dk, code in desc_index.items():
                if key and (key in dk or dk in key):
                    rm_code = code
                    break
        if rm_code is None:
            unresolved.add(mat_name)
            continue
        resolved.append((inter, rm_code, qty))

    print("Resolved to RM_Code:", len(resolved), "/ unresolved:", len(unresolved))
    if unresolved:
        print("  unresolved material names:", sorted(unresolved))

    os.makedirs(out_dir, exist_ok=True)
    with open(os.path.join(out_dir, "bom_from_usage_engineering.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Intermediate", "RM_Code", "RM_Qty_Per_Batch"])
        for inter, rm_code, qty in resolved:
            w.writerow([inter, rm_code, qty])


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "masters"
    )
    extract(src, out)

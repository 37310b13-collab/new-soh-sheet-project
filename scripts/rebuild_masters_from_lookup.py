"""
rm_master.csv・bom.csvを、新しい元データ2ファイルから作り直す。

- Raw Material - Look Up - VFf.xlsx: Catalyst/Slurry/Powder/Solution Data Base の4シート。
  原単位(BOM)と正式なDescription(RM Description列)の情報源。
- Raw materials daily check (Stock シート): 材料一覧・Supplier・Categoryの情報源。
  A列=Material Type(セル結合により空欄行は直前の値を引き継ぐ)、B列=Product(表示名)、
  C列=CSA code(RM_Code。ただしMAT行だけは"Matts"という固定文字列が入っており実際の
  コードではないため、B列の値をRM_Codeとして使う)、D列=Supplier。

突合キーはDescriptionではなくコード(CSA code / RM Code)。詳細な決定事項はユーザーとの
やり取り記録を参照(0/O表記ゆれはO(オー)が正、Corning供給4件は除外、MAT行はChemical
カテゴリとして含める、Consumables(梱包資材)は既存登録済みのEster Film/Original Towel/
PP Filmを除き対象外、71-73行目・85行目のような空欄だけの行は除外)。

使い方:
    python3 scripts/rebuild_masters_from_lookup.py <Look Upファイル> <daily checkファイル> [出力ディレクトリ]
"""
import csv
import os
import re
import sys

import openpyxl

PRESERVED_CODES = {"Ester Film", "Original Towel", "PP Film"}
CODE_FIX = {"0JN": "OJN", "0LJ": "OLJ"}


def norm(s):
    return "".join(ch for ch in str(s).upper() if ch.isalnum())


def load_lookup(path, known_intermediates):
    """Slurry/Powder/Solution/Catalyst Data Baseの4シートから
    (RM_Code -> RM_Description)のインデックスと、BOM行のリストを作る。
    known_intermediatesはintermediate_master.csvのIntermediate列の集合で、
    Catalyst名から抜き出した短縮コードの末尾"s"表記ゆれ(例: "0T420s"→"0T420")を
    解決するために使う。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    desc_by_code = {}
    bom_rows = []

    def harvest_flat(sheet_name):
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            inter, rm_code, rm_desc, qty = row[0], row[3], row[4], row[12]
            if not inter or not rm_code:
                continue
            inter = str(inter).strip()
            rm_code = str(rm_code).strip()
            if rm_desc and rm_code not in desc_by_code:
                desc_by_code[rm_code] = str(rm_desc).strip()
            if isinstance(qty, (int, float)) and qty:
                bom_rows.append((inter, rm_code, float(qty)))

    harvest_flat("Slurry Data Base")
    harvest_flat("Powder Data Base")
    harvest_flat("Solution")

    def catalyst_short_code(name):
        s = name
        if s.startswith("18461-"):
            s = s[len("18461-"):]
        m = re.match(r"[^\s\-]+", s)
        code = m.group(0) if m else s
        if code.startswith("O"):
            code = "0" + code[1:]
        if code not in known_intermediates and code[-1:] in ("s", "S") and code[:-1] in known_intermediates:
            code = code[:-1]
        return code

    ws = wb["Catalyst Data Base"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cat_name, rm_code, rm_desc, qty, sub_sc = row[0], row[3], row[4], row[5], row[7]
        if not sub_sc:
            continue  # Substrate行だけを使う(化学品・スラリー参照行は二重計上になるため除外)
        if rm_desc and "CORNING" in str(rm_desc).upper():
            continue  # Corning供給のSubstrateは対象外(daily check側のCorning除外ルールと揃える)
        sub_sc = str(sub_sc).strip()
        if rm_desc and sub_sc not in desc_by_code:
            desc_by_code[sub_sc] = str(rm_desc).strip()
        code = catalyst_short_code(str(cat_name).strip())
        if isinstance(qty, (int, float)) and qty:
            bom_rows.append((code, sub_sc, float(qty)))

    wb.close()
    return desc_by_code, bom_rows


def load_daily_check(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Stock"]
    records = []
    last_type = None
    for row in ws.iter_rows(min_row=9, max_row=114, values_only=True):
        mtype = row[0] if row[0] else last_type
        if row[0]:
            last_type = row[0]
        product, csa_code, supplier = row[1], row[2], row[3]
        if csa_code is None and supplier is None:
            continue  # 71-73行目・85行目のような空欄だけの残骸行を除外
        records.append({
            "MaterialType": str(mtype).strip() if mtype else "",
            "Product": str(product).strip() if product else "",
            "CSACode": str(csa_code).strip() if csa_code else "",
            "Supplier": str(supplier).strip() if supplier else "",
        })
    wb.close()
    return records


def load_csa_report_descriptions(path):
    """CSA Reportの「Stock invoiced to CSA」シートから、Part No(C列)->Description(D列)の
    索引を作る。MATのようにLook Upの4シートに登場しない材料のDescriptionを補うために使う
    (任意。指定が無ければ何もしない)。"""
    if not path:
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    idx = {}
    ws = wb["Stock invoiced to CSA"]
    for row in ws.iter_rows(min_row=5, values_only=True):
        part_no, desc = row[2], row[3]
        if part_no and desc:
            idx[str(part_no).strip()] = str(desc).strip()
    wb.close()
    return idx


def load_existing_rm_master(path):
    by_code = {}
    if os.path.exists(path):
        with open(path, newline="", encoding="utf-8") as f:
            for r in csv.DictReader(f):
                by_code[r["RM_Code"]] = r
    return by_code


def load_known_intermediates(path):
    known = set()
    if os.path.exists(path):
        with open(path, newline="", encoding="utf-8") as f:
            for r in csv.DictReader(f):
                known.add(r["Intermediate"])
    return known


def main(lookup_path, daily_check_path, out_dir, csa_report_path=None):
    known_intermediates = load_known_intermediates(os.path.join(out_dir, "intermediate_master.csv"))
    desc_by_code, bom_rows = load_lookup(lookup_path, known_intermediates)
    records = load_daily_check(daily_check_path)
    existing = load_existing_rm_master(os.path.join(out_dir, "rm_master.csv"))
    csa_desc_by_code = load_csa_report_descriptions(csa_report_path)

    new_rows = []
    seen_codes = set()
    skipped_corning = []
    skipped_consumables = []

    for rec in records:
        mtype = rec["MaterialType"].upper()
        supplier = rec["Supplier"]
        if "CORNING" in supplier.upper():
            skipped_corning.append(rec["CSACode"])
            continue
        if mtype == "CONSUMABLES":
            skipped_consumables.append(rec["Product"])
            continue  # Ester Film/Original Towel/PP Filmは既存rm_master.csvからそのまま引き継ぐ
        if mtype == "MAT":
            rm_code = rec["Product"]  # MAT行はC列(CSA code)に実コードが入っておらず"Matts"固定のため
            category = "Chemical"
        elif mtype == "CHEMICALS":
            rm_code = rec["CSACode"]
            category = "Chemical"
        elif mtype == "SUBSTRATES":
            rm_code = rec["CSACode"]
            category = "Substrate"
            # daily checkのSupplier列は実際にはメーカー名(NGK等)が入っているが、この
            # ワークブックのSUPPLIER_FILTER=TTAF(TTAF供給材料のみ抽出)はSupplier列で
            # 判定しているため、Substrateは全てTTAF経由の供給として扱いSupplierを
            # "TTAF"に上書きする(以前のビルドで同じ理由で修正済みの挙動を踏襲)。
            supplier = "TTAF"
        else:
            continue
        rm_code = CODE_FIX.get(rm_code, rm_code)
        if not rm_code or rm_code in seen_codes:
            continue
        seen_codes.add(rm_code)

        desc = desc_by_code.get(rm_code) or csa_desc_by_code.get(rm_code) or rec["Product"]
        prior = existing.get(rm_code)
        ttaf_code = prior["TTAF_Code"] if prior else ""
        fixed_qty = prior["FixedWeeklyQty"] if prior else "0"
        # "Hazardous Chemical"はdaily check/Lookupのどちらにも区別が無いため、既存の
        # rm_master.csvで既にこの分類だった材料だけは維持する(それ以外はdaily check側の
        # 分類(Chemical/Substrate)をそのまま採用)。
        if prior and prior["Category"] == "Hazardous Chemical":
            category = "Hazardous Chemical"
        new_rows.append([rm_code, ttaf_code, desc, supplier, category, fixed_qty])

    # 既存rm_master.csvのEster Film/Original Towel/PP Filmをそのまま引き継ぐ
    for code in PRESERVED_CODES:
        if code in existing and code not in seen_codes:
            r = existing[code]
            new_rows.append([r["RM_Code"], r["TTAF_Code"], r["Description"], r["Supplier"],
                              r["Category"], r["FixedWeeklyQty"]])
            seen_codes.add(code)

    rm_master_path = os.path.join(out_dir, "rm_master.csv")
    with open(rm_master_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["RM_Code", "TTAF_Code", "Description", "Supplier", "Category", "FixedWeeklyQty"])
        w.writerows(new_rows)
    print(f"rm_master.csv written: {len(new_rows)} rows")
    print(f"  Corning除外: {len(skipped_corning)} 件 {skipped_corning}")
    print(f"  Consumables除外(Ester Film/Original Towel/PP Film以外): {len(skipped_consumables)} 件")

    old_codes = set(existing.keys())
    dropped = sorted(old_codes - seen_codes)
    added = sorted(seen_codes - old_codes)
    print(f"  旧rm_master.csvにあったが新リストから消えたコード: {len(dropped)} 件 {dropped}")
    print(f"  新規追加されたコード: {len(added)} 件 {added}")

    # ---- bom.csv ----
    valid_codes = seen_codes
    bom_out = []
    unresolved_rm = set()
    for inter, rm_code, qty in bom_rows:
        if rm_code not in valid_codes and rm_code not in desc_by_code:
            unresolved_rm.add(rm_code)
        bom_out.append((inter, rm_code, qty))
    bom_out.sort()

    bom_path = os.path.join(out_dir, "bom.csv")
    with open(bom_path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Intermediate", "RM_Code", "RM_Qty_Per_Batch"])
        for inter, rm_code, qty in bom_out:
            w.writerow([inter, rm_code, qty])
    print(f"bom.csv written: {len(bom_out)} rows")
    if unresolved_rm:
        print(f"  (参考) rm_masterにもLookupのDescription索引にも見つからないRM_Code: {sorted(unresolved_rm)}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    lookup_path = sys.argv[1]
    daily_check_path = sys.argv[2]
    out_dir = sys.argv[3] if len(sys.argv) > 3 else "data/masters"
    csa_report_path = sys.argv[4] if len(sys.argv) > 4 else None
    main(lookup_path, daily_check_path, out_dir, csa_report_path)

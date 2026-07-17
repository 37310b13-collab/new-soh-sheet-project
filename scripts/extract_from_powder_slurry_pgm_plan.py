"""
Powder & Slurry & Pgm Plan (毎月改版, 例: "...- 2026-06-V02 - Delinked.xlsx") から
週次バッチ数(weekly_batches.csv)・原単位の補完データ(bom_from_plan.csv)・substrateマスタ
(substrate_master.csv)を抽出する。

このファイルは~40枚の「化学原料ごとのシート」(例: "AS-200") と、
~3枚の「substrateごとのシート」(例: "Japan GPF Substr")を持つ。

化学原料シート:
  row1: 月ラベル, row2: 材料名 + 週初日(日付, col D以降), row3: 材料コード(CHEM-xxxx) + 週番号,
  row4以降: [中間体名, "No. of batches", 週次バッチ数...] と
            [(空欄), "Usage per batch in kg", 1バッチ使用量(kg), 週次使用量...] の繰り返し。

substrateシート（1シート内に複数substrateのブロックが縦に並ぶ）:
  各ブロック: row(n)  : substrateコード(SSコード, 例 "0JN") + 週初日(日付, col D以降)
              row(n+1): 説明 + TTAF品番 + 週番号
              row(n+2): 完成品コード(Cat, 例 "F2090") + "No. of batches" + 週次バッチ数(=受注数量)...
              row(n+3): "Usage per day" + 1個あたり使用量 + 週次使用量...
  完成品コード(Cat)は原単位展開の「中間体」に相当する役割を果たす
  （中間体を経由せず、完成品の受注数量がそのままsubstrateの使用量になる）。

使い方:
    python3 scripts/extract_from_powder_slurry_pgm_plan.py <入力xlsxパス> [出力ディレクトリ]

出力先は既定で data/masters/ (weekly_batches.csv, bom_from_plan.csv, substrate_master.csv を上書き)。
毎月新しいバージョンのファイルが発行されたら、このスクリプトを再実行して
scripts/build_soh.py を再度動かすことで、SOH_Master.xlsx に反映される。
"""
import openpyxl, csv, datetime, sys, os

RATE_LABELS = ("usage per batch in kg", "usage per day")


def normalize_inter(raw, canonical, canonical_upper):
    raw = raw.strip()
    if raw in canonical:
        return raw
    if raw.upper() in canonical_upper:
        return canonical_upper[raw.upper()]
    cand = "SOL-" + raw.upper()
    if cand in canonical_upper:
        return canonical_upper[cand]
    return raw


def looks_like_code(s):
    if not isinstance(s, str):
        return False
    s = s.strip()
    if not s or len(s) > 12:
        return False
    return any(ch.isalnum() for ch in s)


def extract(src_path, out_dir):
    canonical = set()
    inter_master_path = os.path.join(out_dir, "intermediate_master.csv")
    if os.path.exists(inter_master_path):
        with open(inter_master_path) as f:
            for r in csv.DictReader(f):
                canonical.add(r["Intermediate"])
    canonical_upper = {c.upper(): c for c in canonical}

    wb = openpyxl.load_workbook(src_path, read_only=True, data_only=True)

    batches = {}
    rates = {}
    mat_codes = {}
    substrate_master = {}  # SS_code -> description
    unmapped = set()

    for sn in wb.sheetnames:
        ws = wb[sn]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 5:
            continue

        # 化学原料シート(材料コードがrow3, シート全体で1材料)か、substrateシート
        # (材料コードがrow2, シート内に複数ブロック)かを、まずシート全体でざっくり判定
        row2, row3 = rows[1], rows[2]
        top_date_cols = {ci: v.date() for ci, v in enumerate(row2) if ci >= 3 and isinstance(v, datetime.datetime)}
        top_mat_code = row3[1] if len(row3) > 1 else None
        is_chem_sheet = len(top_date_cols) >= 3 and isinstance(top_mat_code, str) and top_mat_code.startswith("CHEM")

        if is_chem_sheet:
            mat_codes[sn] = top_mat_code
            _harvest_block(rows, 0, top_date_cols, top_mat_code, canonical, canonical_upper,
                            batches, rates, unmapped)
            continue

        # substrateシート: ブロックごとにヘッダー(コード+日付行)を探す
        ri = 0
        while ri < len(rows) - 3:
            r = rows[ri]
            date_cols = {ci: v.date() for ci, v in enumerate(r) if ci >= 3 and isinstance(v, datetime.datetime)}
            code = r[1] if len(r) > 1 else None
            if len(date_cols) >= 3 and looks_like_code(code) and not (isinstance(code, str) and code.startswith("CHEM")):
                desc_row = rows[ri + 1] if ri + 1 < len(rows) else None
                desc = desc_row[1] if desc_row and len(desc_row) > 1 else ""
                substrate_master[code] = desc if isinstance(desc, str) else ""
                mat_codes[f"{sn}:{code}"] = code
                consumed = _harvest_block(rows, ri, date_cols, code, canonical, canonical_upper,
                                           batches, rates, unmapped, rate_labels=RATE_LABELS,
                                           data_offset=2, fallback_to_mat_code=True)
                ri += max(consumed, 1)
            else:
                ri += 1

    print("Material/substrate blocks used:", len(mat_codes))
    print("Unique (Intermediate,Week) batches:", len(batches))
    print("Unique (Intermediate,RM_Code) rates:", len(rates))
    print("Substrate codes found:", len(substrate_master))
    print("Unmapped intermediate/product names (kept as-is):", sorted(unmapped))

    os.makedirs(out_dir, exist_ok=True)
    with open(os.path.join(out_dir, "weekly_batches.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Intermediate", "WeekStart", "Batches"])
        for (inter, wk), qty in sorted(batches.items()):
            if inter in ("-", "", "No of times used"):
                continue
            w.writerow([inter, wk.isoformat(), qty])

    with open(os.path.join(out_dir, "bom_from_plan.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Intermediate", "RM_Code", "RM_Qty_Per_Batch"])
        for (inter, rm), rate in sorted(rates.items()):
            if inter in ("-", "", "No of times used"):
                continue
            w.writerow([inter, rm, rate])

    with open(os.path.join(out_dir, "substrate_master.csv"), "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["RM_Code", "Description"])
        for code, desc in sorted(substrate_master.items()):
            w.writerow([code, desc])


def _harvest_block(rows, start_ri, date_cols, mat_code, canonical, canonical_upper,
                    batches, rates, unmapped, rate_labels=("usage per batch in kg",),
                    data_offset=3, fallback_to_mat_code=False):
    """start_ri行目のコード/日付ヘッダーから始まる1ブロックを読み取る。
    次のブロック(または新しい日付ヘッダー行)に到達したら止まり、消費した行数を返す。

    data_offset: ヘッダー行数(データ行が始まるまでのオフセット)。
    化学原料シート(row1:月ラベル,row2:材料名+日付,row3:コード+週番号,row4~:データ)は3。
    substrateブロック(row(n):コード+日付,row(n+1):説明+週番号,row(n+2)~:データ)は2。

    fallback_to_mat_code: "No. of batches"行の品番(col B)が空欄の場合に、
    そのブロック自身のコード(mat_code)を紐付け先として使う。Ester Film/PP Filmのように
    特定のCat(完成品)コードに紐付かず、フィルム自体が1品目として扱われるシート向け。
    化学原料シート(複数の中間体が同一シート内に並ぶ)では誤結合を避けるためFalseのまま。"""
    current_inter = None
    ri = start_ri + data_offset
    consumed = data_offset
    while ri < len(rows):
        r = rows[ri]
        if len(r) < 2:
            break
        # 次のブロックのヘッダー(コード+日付)に到達したら終了
        next_date_cols = {ci: v.date() for ci, v in enumerate(r) if ci >= 3 and isinstance(v, datetime.datetime)}
        if len(next_date_cols) >= 3 and looks_like_code(r[1]) and ri != start_ri:
            break
        label = r[2] if len(r) > 2 else None
        if isinstance(label, str) and label.strip().lower().startswith("no. of batches"):
            raw_inter = r[1]
            if not raw_inter or not isinstance(raw_inter, str):
                raw_inter = mat_code if fallback_to_mat_code else None
            if not raw_inter or not isinstance(raw_inter, str):
                current_inter = None
            else:
                current_inter = normalize_inter(raw_inter, canonical, canonical_upper)
                if current_inter not in canonical:
                    unmapped.add(raw_inter)
                for ci, wk_date in date_cols.items():
                    if ci >= len(r):
                        continue
                    val = r[ci] or 0
                    key = (current_inter, wk_date)
                    if key not in batches:
                        batches[key] = float(val)
        elif isinstance(r[1], str) and r[1].strip().lower() in rate_labels and current_inter:
            rate = r[2] if len(r) > 2 else None
            if isinstance(rate, (int, float)):
                rates[(current_inter, mat_code)] = float(rate)
        elif (isinstance(label, str) and label.strip().lower() == "total weekly usage") or \
             (isinstance(r[1], str) and r[1].strip().lower() == "total weekly usage"):
            # ブロックの終わり付近の集計行。ここでいったんブロック終了とみなす
            ri += 1
            consumed += 1
            break
        ri += 1
        consumed += 1
    return consumed


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data", "masters"
    )
    extract(src, out)

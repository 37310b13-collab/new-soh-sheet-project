import csv, datetime, sys
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule

import os
REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXTRACTED = os.path.join(REPO_ROOT, "data", "masters") + os.sep

N_WEEKS = int(sys.argv[1]) if len(sys.argv) > 1 else 104
OUT_PATH = sys.argv[2] if len(sys.argv) > 2 else os.path.join(REPO_ROOT, "SOH_Master.xlsx")
ANCHOR_YEAR = int(sys.argv[3]) if len(sys.argv) > 3 else 2026
# サプライヤー別ファイル(例: TTAF専用)を作る場合に指定。M_RawMaterials[Supplier]がこの値と
# 一致する原材料だけに絞り込み、それを1つでも使う中間体だけを連動して残す。
SUPPLIER_FILTER = sys.argv[4] if len(sys.argv) > 4 else None


def monday_containing_jan1(year):
    jan1 = datetime.date(year, 1, 1)
    return jan1 - datetime.timedelta(days=jan1.weekday())


def week_year_and_number(week_start):
    # Week1 of year Y = the Mon-Sun week that contains Jan 1 of Y.
    for y in (week_start.year - 1, week_start.year, week_start.year + 1):
        b = monday_containing_jan1(y)
        nb = monday_containing_jan1(y + 1)
        if b <= week_start < nb:
            return y, (week_start - b).days // 7 + 1
    raise ValueError(f"could not resolve week-year for {week_start}")


START_MONDAY = monday_containing_jan1(ANCHOR_YEAR)  # Week1 start = Monday of the week containing Jan 1


def date_to_week_index(d, n_weeks):
    idx = (d - START_MONDAY).days // 7 + 1
    return max(1, min(n_weeks, idx))

def load_csv(name):
    with open(EXTRACTED + name, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))

rm_master = load_csv("rm_master.csv")
inter_master = load_csv("intermediate_master.csv")
bom = load_csv("bom.csv")
prodmap = load_csv("product_intermediate_map.csv")


def load_csv_optional(name):
    path = os.path.join(EXTRACTED, name)
    if not os.path.exists(path):
        return []
    return load_csv(name)


weekly_batches = load_csv("weekly_batches.csv")

# clean numeric fields, drop rows with missing numeric qty
def to_float(x, default=0.0):
    try:
        return float(x)
    except (TypeError, ValueError):
        return default

# bom.csv (source: Raw Material - Look Up's Slurry/Powder/Solution/Catalyst Data Base sheets)
# already stores the direct per-batch (or per-catalyst-unit) kg amount -- no further
# ratio x batch-size multiplication needed.
for r in bom:
    r["RM_Total_Per_Batch"] = to_float(r["RM_Qty_Per_Batch"])

# dedupe rm_master by RM_Code (defensive)
seen = set()
rm_master_dedup = []
for r in rm_master:
    if r["RM_Code"] in seen:
        continue
    seen.add(r["RM_Code"])
    rm_master_dedup.append(r)
rm_master = rm_master_dedup

seen = set()
inter_dedup = []
for r in inter_master:
    if r["Intermediate"] in seen:
        continue
    seen.add(r["Intermediate"])
    inter_dedup.append(r)
inter_master = inter_dedup

# サプライヤー別ファイル用のフィルタ。M_RawMaterials・M_BOM・M_Intermediates・M_ProductMap・
# PP_Grid・T_Shipments(rm_rowで自動的に絞られる)・Dashboard・Material_Detail・Grid_*・
# PO_Draft_*まで、すべてこの3つのリストを起点に生成されるため、ここで絞り込むだけで
# 以降のシート生成コードは一切変更せずに連動して絞り込まれる。
if SUPPLIER_FILTER:
    _all_inter_names_pre_filter = {r["Intermediate"] for r in inter_master}
    rm_master = [r for r in rm_master if r["Supplier"].strip() == SUPPLIER_FILTER]
    _filt_codes = {r["RM_Code"].upper() for r in rm_master}
    # bom.csvのRM_Codeは、購入する原材料(rm_master由来)だけでなく、他の中間体を経由専用で
    # 消費する行(例: TSP-618がTSZ-616を使う)も含む。後者はTSZ-616自身が購入品ではなく
    # rm_masterに載らないため、rm_masterのコード集合だけでフィルタすると誤って除外され、
    # PP_Gridの経由専用中間体(数式によるパススルー)の元になる行が消えてしまう。
    # そのためRM_Codeが「他の中間体名」である行はサプライヤーに関係なく常に残す。
    bom = [r for r in bom
           if r["RM_Code"].upper() in _filt_codes or r["RM_Code"] in _all_inter_names_pre_filter]
    _filt_inter_names = {r["Intermediate"] for r in bom}
    inter_master = [r for r in inter_master if r["Intermediate"] in _filt_inter_names]
    prodmap = [r for r in prodmap if r["Intermediate"] in _filt_inter_names]
    print(f"[SUPPLIER_FILTER={SUPPLIER_FILTER}] RM: {len(rm_master)}, "
          f"Intermediates: {len(inter_master)}, BOM rows: {len(bom)}")

# rm_masterの並び順は、Dashboard・Material_Detail・M_RawMaterials・T_SelfStock・T_TTAFStock・
# PO_Draft_*等、この行順を起点に生成される全シートに反映される。要望に基づき
# 「Substrate → その他Chemical → Ester Film・PP Film → TPZ系」の順に並べる
# (Python標準のsorted()は安定ソートなので、各グループ内は元の順序を保ったまま並び替わる)。
_ESTER_PP = {"ESTER FILM", "PP FILM"}
def _rm_sort_group(r):
    code = r["RM_Code"].strip().upper()
    desc = r["Description"].strip().upper()
    is_tpz = "TPZ" in desc or "TZP" in desc  # LookupのRM Descriptionに"TZP"表記の誤記があるため両対応
    if code in _ESTER_PP:
        return 2
    if is_tpz:
        return 3
    if r["Category"] == "Substrate":
        return 0
    return 1
rm_master.sort(key=_rm_sort_group)

# weekly batch counts, source: Powder & Slurry & Pgm Plan ("No. of batches" rows, one entry
# per intermediate per week, extracted from ~36 per-material sheets). Keyed by actual calendar
# date (WeekStart) so it lines up correctly regardless of ANCHOR_YEAR.
plan_lookup = {}
for r in weekly_batches:
    ws_str = r.get("WeekStart")
    if not ws_str:
        continue
    try:
        wk_date = datetime.date.fromisoformat(ws_str)
    except ValueError:
        continue
    qty = to_float(r["Batches"], 0.0)
    plan_lookup.setdefault(r["Intermediate"], {})[wk_date] = qty

print(f"RM master: {len(rm_master)}, Intermediates: {len(inter_master)}, BOM rows: {len(bom)}, ProductMap: {len(prodmap)}")

# ---------------------------------------------------------------------------
wb = openpyxl.Workbook()
wb.remove(wb.active)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

def add_table(ws, name, ref, style="TableStyleMedium9"):
    t = Table(displayName=name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name=style, showRowStripes=True)
    ws.add_table(t)

# ============================================================ Cal_Weeks
# WeekIndex(内部の連番, グリッドの列位置に使用) と、年ごとに1へリセットするWeekOfYear/Label(表示用)。
# Week1 = AnchorYearの1月1日を含む月〜日の週。日付・週番号はすべてExcelの数式で計算される
# （B1のAnchorYearを変更すれば全体が再計算される）。
week_labels = {}  # Python側でも列位置の対応付けに使うため、値は事前計算して保持しておく
CAL_HEADER_ROW = 3  # row1=AnchorYear入力, row2=列見出し, row3以降=データ
ws = wb.create_sheet("Cal_Weeks")
ws["A1"] = "AnchorYear"
ws["B1"] = ANCHOR_YEAR
ws["A1"].font = Font(bold=True)
ws["B1"].fill = INPUT_FILL
ws.append([])
ws.append(["WeekIndex", "WeekStart", "Year", "WeekOfYear", "Label", "WeekEnd", "MonthYearLabel"])
for i in range(1, N_WEEKS + 1):
    r = CAL_HEADER_ROW + i
    if i == 1:
        wk_start_formula = "=DATE($B$1,1,1)-WEEKDAY(DATE($B$1,1,1),3)"
    else:
        wk_start_formula = f"=B{r-1}+7"
    year_formula = (
        f"=IF(B{r}>=DATE(YEAR(B{r})+1,1,1)-WEEKDAY(DATE(YEAR(B{r})+1,1,1),3),YEAR(B{r})+1,"
        f"IF(B{r}>=DATE(YEAR(B{r}),1,1)-WEEKDAY(DATE(YEAR(B{r}),1,1),3),YEAR(B{r}),YEAR(B{r})-1))"
    )
    week_of_year_formula = f"=(B{r}-(DATE(C{r},1,1)-WEEKDAY(DATE(C{r},1,1),3)))/7+1"
    label_formula = f'=C{r}&"-W"&TEXT(D{r},"00")'
    week_end_formula = f"=B{r}+6"
    month_year_formula = f'=TEXT(B{r},"mmm-yy")'
    ws.append([i, wk_start_formula, year_formula, week_of_year_formula, label_formula, week_end_formula, month_year_formula])
    ws.cell(row=r, column=2).number_format = "yyyy-mm-dd"
    ws.cell(row=r, column=6).number_format = "yyyy-mm-dd"
    # Python側の記録用(他シートのビルド時の列位置計算・MATCH対象文字列の生成に使用。
    # 実際のセルの値は上記の通りExcel数式で計算される)
    wk_start = START_MONDAY + datetime.timedelta(weeks=i - 1)
    yr, wn = week_year_and_number(wk_start)
    week_labels[i] = f"{yr}-W{wn:02d}"
style_header(ws, 7, row=CAL_HEADER_ROW)
add_table(ws, "Cal_Weeks", f"A{CAL_HEADER_ROW}:G{CAL_HEADER_ROW+N_WEEKS}")
for col, w in zip("ABCDEFG", [10, 12, 8, 11, 12, 12, 14]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = f"A{CAL_HEADER_ROW+1}"

# ============================================================ M_RawMaterials
# 基準在庫は下限・上限の2つを持つ(Dashboardの週次の赤/緑/青のハイライトに使用)。
# 以前の単一しきい値のSafetyStock_Qty_要入力を、下限・上限の2列に置き換えた。
# 固定週次消費量_要入力: BOM(M_BOM/PP_Grid)経由の消費量計算とは別に、材料によっては
# 毎週固定量を消費する(例: Original Towelは梱包資材で、生産中間体の構成とは無関係に
# 週200枚など一定量を消費する)。Grid_Requirementはこの列の値をBOMベースの計算結果に
# 単純に加算する。0の材料には影響しない。ここはExcel上のセルなので、消費量が変わった際も
# 再生成なしにそのまま書き換えるだけで反映される。
ws = wb.create_sheet("M_RawMaterials")
ws.append(["Part Name", "Description", "Supplier", "Category", "UOM",
           "基準在庫下限_要入力", "基準在庫上限_要入力", "LeadTime_Weeks_要入力", "TTAF_Code",
           "固定週次消費量_要入力"])
for r in rm_master:
    ws.append([r["RM_Code"], r["Description"], r["Supplier"], r["Category"], "kg", 0, 0, 4,
               r.get("TTAF_Code", ""), float(r.get("FixedWeeklyQty", 0) or 0)])
n = len(rm_master) + 1
style_header(ws, 10)
add_table(ws, "M_RawMaterials", f"A1:J{n}")
for col, w in zip("ABCDEFGHIJ", [14, 34, 14, 16, 8, 18, 18, 18, 16, 20]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

# ============================================================ M_Intermediates
ws = wb.create_sheet("M_Intermediates")
ws.append(["Intermediate", "Type", "PGM", "BatchSize"])
for r in inter_master:
    ws.append([r["Intermediate"], r["Type"], r["PGM"], to_float(r["Batch_Size"], 1.0)])
n = len(inter_master) + 1
style_header(ws, 4)
add_table(ws, "M_Intermediates", f"A1:D{n}")
for col, w in zip("ABCD", [14, 10, 8, 12]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

# ============================================================ M_BOM
# D列(PPGridRow)は、Grid_Requirementが毎週×毎材料でPP_Grid内をMATCHし直す(重い)代わりに
# 「このIntermediateがPP_Grid内の何行目か」を1回だけ計算しておく内部ヘルパー列。
# VBAでM_BOMに行を追加しても、Excelのテーブル機能が数式列を自動で複製するため引き続き機能する。
ws = wb.create_sheet("M_BOM")
ws.append(["Intermediate", "Part Name", "RM_Qty_Per_Batch", "PPGridRow"])
for i, r in enumerate(bom):
    rr = i + 2
    ws.append([r["Intermediate"], r["RM_Code"], r["RM_Total_Per_Batch"],
               f'=IFERROR(MATCH($A{rr},PP_Grid[Intermediate],0),99999)'])
n = len(bom) + 1
style_header(ws, 4)
add_table(ws, "M_BOM", f"A1:D{n}")
for col, w in zip("ABCD", [14, 12, 18, 12]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

# ============================================================ M_ProductMap
ws = wb.create_sheet("M_ProductMap")
ws.append(["Intermediate", "Product_Code"])
for r in prodmap:
    ws.append([r["Intermediate"], r["Product_Code"]])
n = len(prodmap) + 1
style_header(ws, 2)
add_table(ws, "M_ProductMap", f"A1:B{n}")
for col, w in zip("AB", [14, 16]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"

rm_row = {r["RM_Code"]: i + 2 for i, r in enumerate(rm_master)}          # Grid_* row for RM_Code (row1=header)
# 外部データ(shipments_all.csv等)がRM_Codeを大文字/小文字違いで記載している場合があるため、
# 「このRM_Codeがrm_masterに存在するか」の判定は大文字小文字を区別せずに行う
# (Excelの数式側の比較はもともと大文字小文字を区別しないため、Python側もそれに合わせる)。
# 一致したら、rm_master側の正しい表記(canonical case)に揃えて使う。
rm_code_canonical = {code.upper(): code for code in rm_row}
def normalize_rm_code(code):
    if not code:
        return None
    return rm_code_canonical.get(code.upper())
inter_row = {r["Intermediate"]: i + 2 for i, r in enumerate(inter_master)}  # PP_Grid row for Intermediate

# T_SelfStock/T_TTAFStock(材料×週のグリッド)の行位置。1行目はDashboard/Material_Detailと
# 同様に選択週の入力欄(C1)用に空けておき、4段見出し(月-年/日付/週No/週ラベル)はその下から、
# 材料はさらにその下から並ぶ。
SS_MONTHYEAR_ROW, SS_DATE_ROW, SS_WEEKNO_ROW, SS_TABLE_ROW = 2, 3, 4, 5
SS_DATA_START_ROW = SS_TABLE_ROW + 1
ss_row_map = {r["RM_Code"]: i + SS_DATA_START_ROW for i, r in enumerate(rm_master)}  # T_SelfStock/T_TTAFStock row

def week_col(week_idx):
    return get_column_letter(1 + week_idx)  # col A = label, col B = week1

# ============================================================ PP_Grid (production plan, INPUT)
ws = wb.create_sheet("PP_Grid")
header = ["Intermediate"] + [week_labels[w] for w in range(1, N_WEEKS + 1)]
ws.append(header)
for i, r in enumerate(inter_master):
    row_num = i + 2
    inter = r["Intermediate"]
    row_vals = [inter]
    weekvals = plan_lookup.get(inter, {})
    for w in range(1, N_WEEKS + 1):
        wk_date = START_MONDAY + datetime.timedelta(weeks=w - 1)
        row_vals.append(to_float(weekvals.get(wk_date, 0), 0))
    ws.append(row_vals)
    for w in range(1, N_WEEKS + 1):
        ws.cell(row=row_num, column=1 + w).fill = INPUT_FILL
n = len(inter_master) + 1
style_header(ws, N_WEEKS + 1)
add_table(ws, "PP_Grid", f"A1:{week_col(N_WEEKS)}{n}")
ws.column_dimensions["A"].width = 14
for w in range(1, N_WEEKS + 1):
    ws.column_dimensions[week_col(w)].width = 7
ws.freeze_panes = "B2"

# ---- 中間体が別の中間体の原料になっているケース(例: TSZ-616はTSP-618等のスラリーの
# 原料になっているが、TSZ-616自身は購入品ではなく、「Powder & Slurry & Pgm Plan」にも
# 独自の週次バッチ計画を持たない)への対応。このような「経由専用」の中間体は、PP_Grid上の
# セルを固定値ではなく、それを使う側(TSP-618等、実バッチ数を持つ中間体)からの需要を
# Grid_Requirementと全く同じSUMPRODUCTパターンで逆算する数式にする。これによりCHEM-1600等の
# 週次所要量が正しく計算される一方、TSZ-616自身はrm_masterに載らないためDashboard等には
# 一切表示されない。Python側で一度だけ数式を書き込むだけで、以降はExcelの数式として
# 完結する(TSZ-616のレシピや、それを使うスラリーの増減があっても再生成不要で自動追従する)。
plan_lookup_names = set(plan_lookup.keys())
bom_rm_code_names = {b["RM_Code"] for b in bom}
inter_master_names = {r["Intermediate"] for r in inter_master}
# Ester Film/PP Filmのような自己参照BOM行(Intermediate==RM_Code。1バッチ=1個消費という
# 簡易表現のためのもの)は、そのままパススルー扱いにすると「自分自身のPP_Grid行を参照する
# 数式」になり循環参照になってしまうため、対象から除外する(独自の週次データが無ければ
# 素直に0のままにする。RefreshWeeklyBatches等で実データが入れば通常通り使われる)。
self_referencing_names = {b["Intermediate"] for b in bom if b["Intermediate"] == b["RM_Code"]}
passthrough_intermediates = sorted(
    (bom_rm_code_names & inter_master_names) - plan_lookup_names - self_referencing_names
)
if passthrough_intermediates:
    print("PP_Grid: pass-through(数式)中間体:", passthrough_intermediates)
inter_batch_size = {r["Intermediate"]: to_float(r["Batch_Size"], 1.0) for r in inter_master}
for name in passthrough_intermediates:
    trow = inter_row[name]
    batch_size = inter_batch_size.get(name, 1.0) or 1.0
    for w in range(1, N_WEEKS + 1):
        ws.cell(row=trow, column=1 + w).value = (
            f"=SUMPRODUCT((M_BOM[Part Name]=$A{trow})*M_BOM[RM_Qty_Per_Batch]*"
            f"IFERROR(INDEX(PP_Grid[#Data],M_BOM[PPGridRow],{w + 1}),0))/{batch_size}"
        )
        ws.cell(row=trow, column=1 + w).fill = PatternFill(fill_type=None)  # INPUT_FILLを解除(数式セルのため)

# ============================================================ T_OpeningStock (INPUT)
ws = wb.create_sheet("T_OpeningStock")
ws.append(["Part Name", "Opening_Qty_要入力", "AsOf"])
for r in rm_master:
    ws.append([r["RM_Code"], 0, START_MONDAY])
n = len(rm_master) + 1
style_header(ws, 3)
add_table(ws, "T_OpeningStock", f"A1:C{n}")
for c in range(2, 4):
    for row in range(2, n + 1):
        ws.cell(row=row, column=c).fill = INPUT_FILL
for col, w in zip("ABC", [14, 16, 12]):
    ws.column_dimensions[col].width = w

# WeekIndexを「日付から毎回ライブ計算する式」にするための共通部品。
# Cal_Weeks!$B$1(AnchorYear)が変わっても、記録した日付から正しい週番号を再計算できるため、
# AnchorYearを進めても過去の実績データ(T_Shipments/T_StockCount/T_SelfStock/T_TTAFStock)が
# 別の週のものとして誤表示されることがなくなる。
#
# 2種類の式を使い分ける:
#   week_index_formula_clamped: 表示ウィンドウの外側の日付は端の週(1 or N_WEEKS)に寄せる。
#     T_Shipments(まだ着荷していない発注)用。古い発注でも「本来もう届いているはず」として
#     week1に表示され続けてほしいため。
#   week_index_formula_strict : 表示ウィンドウの外側の日付はどの週にも一致させず空欄にする。
#     T_StockCount/T_SelfStock/T_TTAFStock(実績の記録)用。クランプしてしまうと、AnchorYearを
#     進めた後にウィンドウ外へ出た古い実績データがweek1の集計に紛れ込み、直近の実績値が
#     狂ってしまうため。
ANCHOR_MONDAY_EXPR = "(DATE(Cal_Weeks!$B$1,1,1)-WEEKDAY(DATE(Cal_Weeks!$B$1,1,1),3))"
def week_index_formula_clamped(date_cell_ref):
    return f'=IFERROR(MAX(1,MIN({N_WEEKS},INT(({date_cell_ref}-{ANCHOR_MONDAY_EXPR})/7)+1)),"")'
def week_index_formula_strict(date_cell_ref):
    raw = f'(INT(({date_cell_ref}-{ANCHOR_MONDAY_EXPR})/7)+1)'
    return f'=IFERROR(IF(AND({raw}>=1,{raw}<={N_WEEKS}),{raw},""),"")'

# ============================================================ T_Shipments (INPUT, seeded from Shipping Schedule)
raw_shipments = load_csv("shipments_all.csv")
ship_rows = []
for row in raw_shipments:
    rm_code = normalize_rm_code(row["RM_Code"])
    if rm_code is None:
        continue
    eta_str = row["Latest_ETA"]
    if not eta_str:
        continue
    eta = datetime.date.fromisoformat(eta_str)
    recv_str = row["Received_Date"]
    recv_date = datetime.date.fromisoformat(recv_str) if recv_str else None
    # Only keep shipments relevant to the forecast horizon: still pending, or
    # received recently (last 14d) / due soon. Older fully-received shipments
    # are assumed already folded into T_OpeningStock and must not be double counted.
    horizon_end = START_MONDAY + datetime.timedelta(weeks=N_WEEKS)
    relevant = (recv_date is None and eta < horizon_end) or \
               (recv_date is not None and recv_date >= START_MONDAY - datetime.timedelta(days=14) and recv_date < horizon_end) or \
               (recv_date is None and eta >= START_MONDAY - datetime.timedelta(days=14))
    if not relevant:
        continue
    ship_rows.append([rm_code, row["PO_No"] or "", None, to_float(row["Confirmed_Qty"], 0), eta, recv_date, row["Status"] or ""])
    # ship_rows layout: RM_Code, PO_No, Order_Date(unknown from source, left blank for manual entry), Confirmed_Qty, Latest_ETA, Received_Date, Status

ws = wb.create_sheet("T_Shipments")
ws.append(["Part Name", "PO_No", "Order_Date_発注日", "Confirmed_Qty", "Latest_ETA", "Received_Date", "Status",
           "Effective_Week", "Order_Month"])
if not ship_rows:
    # Excelのテーブル機能は見出し行のみ(データ0行)の範囲を許容しないため、
    # 該当する発注が無い場合はダミー行を1行入れておく（要削除・上書き可）。
    ship_rows = [["(例) CHEM-1010", "", None, 0, START_MONDAY, None, ""]]
start_row = 2
for i, r in enumerate(ship_rows):
    row_num = start_row + i
    ws.append(r + [None, None])
    ws.cell(row=row_num, column=3).fill = INPUT_FILL  # Order_Date is not in the source file; input by hand
    # Effective_Week: use Received_Date if present else Latest_ETA; week index via anchor arithmetic, clamped to sheet horizon.
    # 起点日はCal_Weeks!$B$1(AnchorYear)からその都度計算する(week1の起点と同じ式、week_index_formula_clamped参照)。
    # 以前はビルド時点のAnchorYearを固定値としてDATE(...)に埋め込んでいたため、
    # AnchorYearをシート上で変更してもこの列だけ古い年のままズレる不具合があった。
    ws.cell(row=row_num, column=8).value = week_index_formula_clamped(
        f'IF(F{row_num}="",E{row_num},F{row_num})'
    )
    ws.cell(row=row_num, column=9).number_format = "yyyy-mm"
n = len(ship_rows) + 1
style_header(ws, 9)
add_table(ws, "T_Shipments", f"A1:I{n}")
for col, w in zip("ABCDEFGHI", [12, 14, 16, 14, 14, 14, 14, 14, 12]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A2"
print("Shipment rows seeded:", len(ship_rows))

# T_PlannedOrders(計画中の発注の入力シート)は廃止しました。発注数量はMaterial_Detailの
# 「Order(発注予定,kg)」行に直接手入力する方式に変更したため、このシートの役割は不要に
# なりました(5.10章参照)。
ws.freeze_panes = "A2"

# ============================================================ T_StockCount (INPUT, physical count overrides)
# 列順: RM_Code, Date(手入力=棚卸を実施した日), WeekIndex(Dateから自動計算), CountedQty, Notes
ws = wb.create_sheet("T_StockCount")
ws.append(["Part Name", "Date_棚卸実施日", "WeekIndex", "CountedQty", "Notes"])
ws.append(["(例) CHEM-1010", START_MONDAY, None, 0, "棚卸実施時にこの行へ追記(Dateを入力するとWeekIndexは自動計算されます)"])
ws.cell(row=2, column=3).value = week_index_formula_strict("$B2")
style_header(ws, 5)
add_table(ws, "T_StockCount", "A1:E2")
ws.cell(row=2, column=2).number_format = "yyyy-mm-dd"
for col, w in zip("ABCDE", [16, 14, 10, 12, 30]):
    ws.column_dimensions[col].width = w

# ============================================================ T_SelfStock / T_TTAFStock (自社/TTAF倉庫の実績)
# 設計: VBAが書き込む生データは「実施日」をキーにした安全な形(_Logシート、非表示)に保存する。
# WeekIndexは各行のDateから毎回ライブ計算されるため、AnchorYearを何度・どんな頻度で進めても
# 実績データが「別の週のもの」として誤表示されることはない。
# 目に見える方のT_SelfStock/T_TTAFStockシートは、Dashboardと同じ「材料×週」のグリッド形式にし、
# 値は一切保存せず_Logシートを参照するSUMIFSだけで組み立てる(常に最新計算)。こうすることで、
# グリッドの列(週)が現在のAnchorYearに紐づいていても、実データ自体は_Logシート側で安全に
# 保たれたままなので、AnchorYearを何度動かしても実績の生データが壊れることはない。
# （_Logシート側は、同じ週内に複数回VBAで取り込んでも1行に上書きされる。以前は実施日そのもの
# をキーにしていたため、日次で取り込むたびに行が積み上がっていたが、それをその週の月曜日
# （実際の暦日から計算。AnchorYearには依存しない）をキーにする方式に変更して解消している）。
def build_actual_stock_sheets(name, qty_col, sample_rows):
    log_name = f"{name}_Log"

    # ---- 非表示の生ログ(VBAがここに書き込む。実施日ベースで安全) ----
    ws_log = wb.create_sheet(log_name)
    ws_log.append(["Part Name", "Date", "WeekIndex", qty_col])
    rows_written = 0
    for r in sample_rows:
        d = datetime.date.fromisoformat(r["Date"])
        ws_log.append([r["RM_Code"], d, None, to_float(r[qty_col], 0)])
        rows_written += 1
    if rows_written == 0:
        # Excelのテーブル機能は見出し行のみ(データ0行)の範囲を許容しないための構造上の
        # ダミー行。表示中の104週ウィンドウから確実に外れる日付(WeekIndexが空欄になる)を
        # 使うことで、グリッド側には一切表示されないようにしている(サンプルデータに
        # 見えてしまわないように)。
        ws_log.append(["(ダミー行、削除不可)", datetime.date(2000, 1, 3), None, 0])
        rows_written = 1
    for row_i in range(2, rows_written + 2):
        ws_log.cell(row=row_i, column=3).value = week_index_formula_strict(f"$B{row_i}")
        ws_log.cell(row=row_i, column=2).number_format = "yyyy-mm-dd"
    style_header(ws_log, 4)
    add_table(ws_log, log_name, f"A1:D{rows_written+1}")
    for c, w in zip("ABCD", [16, 12, 10, 12]):
        ws_log.column_dimensions[c].width = w
    print(f"{log_name} rows seeded:", rows_written)

    # ---- 目に見えるグリッド(材料×週。すべて数式で_Logシートから毎回計算、値は保存しない) ----
    ws_grid = wb.create_sheet(name)

    # 選択週の入力欄(C1)。Dashboard/Material_Detailと全く同じ仕組み(SUMPRODUCTで
    # 「現在年×入力した週No」に一致するWeekIndexを求める)。VBAのJumpToSelectedWeekを
    # 導入していれば、入力するとその週列のすぐ右(B列側)に自動でスクロールする。
    ws_grid["A1"] = "選択週を入力（例: W23。現在年の週Noで検索します）"
    ws_grid["A1"].font = Font(bold=True)
    ws_grid["C1"] = ""
    ws_grid["C1"].fill = INPUT_FILL
    ws_grid["C1"].font = Font(bold=True, size=12)
    _ss_cal_first = CAL_HEADER_ROW + 1
    _ss_cal_last = CAL_HEADER_ROW + N_WEEKS
    _ss_cal_year_rng = f"Cal_Weeks!$C${_ss_cal_first}:$C${_ss_cal_last}"
    _ss_cal_weekofyear_rng = f"Cal_Weeks!$D${_ss_cal_first}:$D${_ss_cal_last}"
    _ss_cal_weekindex_rng = f"Cal_Weeks!$A${_ss_cal_first}:$A${_ss_cal_last}"
    _ss_cal_label_rng = f"Cal_Weeks!$E${_ss_cal_first}:$E${_ss_cal_last}"
    _ss_wk_match = (f"({_ss_cal_year_rng}=Cal_Weeks!$B$1)*"
                    f'({_ss_cal_weekofyear_rng}=VALUE(SUBSTITUTE(UPPER(TRIM($C$1)),"W","")))')
    ws_grid["F1"] = (
        f'=IFERROR(IF(SUMPRODUCT({_ss_wk_match})=0,"",SUMPRODUCT({_ss_wk_match}*({_ss_cal_weekindex_rng}))),"")'
    )
    ws_grid["F1"].font = Font(size=8, color="808080")
    ws_grid["E1"] = "→WeekIndex"
    ws_grid["E1"].font = Font(size=8, color="808080")
    ws_grid["D1"] = (
        f'=IF($C$1="","週Noを入力してください（例: W23）",'
        f'IF($F$1="","該当週が見つかりません（今年の週Noか確認してください）",'
        f'INDEX({_ss_cal_label_rng},$F$1)&" が該当週です（VBA導入時は自動でスクロールします）"))'
    )
    ws_grid["D1"].font = Font(bold=True, color="0563C1")

    for w in range(1, N_WEEKS + 1):
        col = 1 + w
        cal_row = CAL_HEADER_ROW + w
        if w == 1:
            my_formula = f"='Cal_Weeks'!G{cal_row}"
        else:
            prev_cal_row = CAL_HEADER_ROW + w - 1
            my_formula = (
                f'=IF(TEXT(\'Cal_Weeks\'!B{cal_row},"mmm-yy")<>TEXT(\'Cal_Weeks\'!B{prev_cal_row},"mmm-yy"),'
                f"'Cal_Weeks'!G{cal_row},\"\")"
            )
        ws_grid.cell(row=SS_MONTHYEAR_ROW, column=col, value=my_formula)
        dcell = ws_grid.cell(row=SS_DATE_ROW, column=col, value=f"='Cal_Weeks'!B{cal_row}")
        dcell.number_format = "m/d"
        ws_grid.cell(row=SS_WEEKNO_ROW, column=col, value=f"='Cal_Weeks'!D{cal_row}")
        ws_grid.cell(row=SS_TABLE_ROW, column=col, value=week_labels[w])
    for r in (SS_MONTHYEAR_ROW, SS_DATE_ROW, SS_WEEKNO_ROW):
        for c in range(1, N_WEEKS + 2):
            ws_grid.cell(row=r, column=c).fill = PatternFill("solid", fgColor="D9E1F2")
            ws_grid.cell(row=r, column=c).font = Font(bold=(r == SS_MONTHYEAR_ROW))
    ws_grid.cell(row=SS_TABLE_ROW, column=1, value="Part Name")

    for i, r in enumerate(rm_master):
        rr = SS_DATA_START_ROW + i
        rm = r["RM_Code"]
        ws_grid.cell(row=rr, column=1, value=rm)
        for w in range(1, N_WEEKS + 1):
            # 記録が無い週は(SUMIFSの0ではなく)空欄"" にする。0という実績が記録された週と
            # 「記録自体が無い」週を区別できるようにするため(Grid_Stockの手動棚卸>実績>
            # ロールフォワードという優先順位判定が、記録の有無を見て動いているため)。
            has_record = f"COUNTIFS({log_name}[Part Name],$A{rr},{log_name}[WeekIndex],{w})"
            ws_grid.cell(row=rr, column=1 + w, value=(
                f'=IF({has_record}=0,"",SUMIFS({log_name}[{qty_col}],{log_name}[Part Name],$A{rr},'
                f'{log_name}[WeekIndex],{w}))'
            ))
    n_last_row = SS_DATA_START_ROW + len(rm_master) - 1
    style_header(ws_grid, N_WEEKS + 1, row=SS_TABLE_ROW)
    # ヘッダーが数式/複数行(月-年・日付・週No)のため、Dashboardと同様にExcelのテーブル機能は
    # 使わず、罫線・縞模様の手動書式で「テーブルらしい」見た目にする。
    thin_ss = Side(style="thin", color="BFBFBF")
    border_ss = Border(left=thin_ss, right=thin_ss, top=thin_ss, bottom=thin_ss)
    for rr2 in range(SS_TABLE_ROW, n_last_row + 1):
        stripe = (rr2 - SS_TABLE_ROW) % 2 == 1
        for c in range(1, N_WEEKS + 2):
            cell = ws_grid.cell(row=rr2, column=c)
            cell.border = border_ss
            if rr2 > SS_TABLE_ROW and stripe:
                cell.fill = PatternFill("solid", fgColor="F2F2F2")
    ws_grid.column_dimensions["A"].width = 14
    for w in range(1, N_WEEKS + 1):
        ws_grid.column_dimensions[week_col(w)].width = 9
    ws_grid.freeze_panes = f"B{SS_DATA_START_ROW}"

    # 選択中の週(F1のWeekIndex)に該当する列を太枠で強調。行数がDashboardと同程度(材料数のみ、
    # Material_Detailのような数千行にはならない)なので、条件付き書式を付けてもパフォーマンス
    # 上の懸念はない。
    ss_week_select_border = Border(left=Side(style="thick", color="BF8F00"),
                                    right=Side(style="thick", color="BF8F00"))
    ws_grid.conditional_formatting.add(
        f"B{SS_TABLE_ROW}:{week_col(N_WEEKS)}{n_last_row}",
        FormulaRule(
            formula=['(COLUMN()-COLUMN($B$1)+1)=$F$1'],
            border=ss_week_select_border,
        )
    )
    print(f"{name} grid built:", len(rm_master), "materials x", N_WEEKS, "weeks")
    return n_last_row

build_actual_stock_sheets("T_SelfStock", "Self_Qty", [])
build_actual_stock_sheets("T_TTAFStock", "TTAF_Qty", [])

# ============================================================ Grid_Requirement / Grid_Incoming / Grid_Stock
# Grid_Requirementは、以前は「BOM行×週」を1行ずつ展開した中間表(Calc_Demand, 73,944行)を
# SUMIFSで週次集計していたが、これが実Excelで開く・編集する・スクロールするたびに
# 極めて重い処理となり、フリーズ・強制終了の主因になっていた（10,504セル×74,000行SUMIFS
# ≈ 15億回超の比較）。M_BOM(711行)とPP_Grid(週次バッチ数)からSUMPRODUCTで直接集計する
# 方式に変更し、中間表を廃止（同じ計算結果を約200分の1の計算量で得られる）。
ws_req = wb.create_sheet("Grid_Requirement")
ws_in = wb.create_sheet("Grid_Incoming")
ws_st = wb.create_sheet("Grid_Stock")
# Grid_TheoreticalStock: Grid_Stockと行位置(rr)を完全にそろえた、常に「前週+入庫-消費」だけで
# 計算する純粋なロールフォワード専用シート。T_StockCount(棚卸)や自社/TTAF実績には一切
# 反応しない(Grid_Stockのような優先順位の切り替えが無い)。Dashboardで「理論在庫」行として
# 表示し、実際の値(Grid_Stock=「実在庫」行)との乖離を確認できるようにするために追加した。
ws_theo = wb.create_sheet("Grid_TheoreticalStock")

header = ["Part Name"] + [week_labels[w] for w in range(1, N_WEEKS + 1)]
for ws_ in (ws_req, ws_in, ws_st, ws_theo):
    ws_.append(header)

for i, r in enumerate(rm_master):
    rr = i + 2
    rm = r["RM_Code"]
    for ws_ in (ws_req, ws_in, ws_st, ws_theo):
        ws_.append([rm] + [None] * N_WEEKS)
    for w in range(1, N_WEEKS + 1):
        cl = week_col(w)
        # M_BOMのうちRM_Code=このRMの行だけを対象に、原単位×その週のバッチ数(PP_GridRow経由で
        # 週ごとのMATCHをせず直接INDEX)を合計する。PP_Grid内の列位置(w+1列目=Intermediate列の次)
        # は週ごとに固定できるため、MATCHは行位置(PPGridRow, M_BOM側で1回だけ計算済み)のみで済む。
        # BOMベースの消費量に加え、M_RawMaterialsの固定週次消費量_要入力(Original Towel等、
        # 生産中間体の構成とは無関係に毎週一定量を消費する梱包資材向け)を単純加算する。
        ws_req.cell(row=rr, column=1 + w).value = (
            f"=SUMPRODUCT((M_BOM[Part Name]=$A{rr})*M_BOM[RM_Qty_Per_Batch]*"
            f"IFERROR(INDEX(PP_Grid[#Data],M_BOM[PPGridRow],{w + 1}),0))"
            f"+IFERROR(INDEX(M_RawMaterials[固定週次消費量_要入力],MATCH($A{rr},M_RawMaterials[Part Name],0)),0)"
        )
        ws_in.cell(row=rr, column=1 + w).value = (
            f"=SUMIFS(T_Shipments[Confirmed_Qty],T_Shipments[Part Name],$A{rr},T_Shipments[Effective_Week],{w})"
        )
        # T_StockCountは棚卸(手動・低頻度)のためCOUNTIFS/SUMIFSのままでよい。
        # T_SelfStock/T_TTAFStockは「材料×週」のグリッド形式(直接セル参照)になったため、
        # SUMIFS不要でGrid_Requirement/Incomingと同じ直接参照で済む(高速・行数増加の心配もない)。
        has_count = f"COUNTIFS(T_StockCount[Part Name],$A{rr},T_StockCount[WeekIndex],{w})"
        count_val = f"SUMIFS(T_StockCount[CountedQty],T_StockCount[Part Name],$A{rr},T_StockCount[WeekIndex],{w})"
        ss_row = ss_row_map[rm]
        has_self = f"('T_SelfStock'!{cl}{ss_row}<>\"\")"
        self_val = f"'T_SelfStock'!{cl}{ss_row}"
        has_ttaf = f"('T_TTAFStock'!{cl}{ss_row}<>\"\")"
        ttaf_val = f"'T_TTAFStock'!{cl}{ss_row}"

        if w == 1:
            prior = f'IFERROR(INDEX(T_OpeningStock[Opening_Qty],MATCH($A{rr},T_OpeningStock[Part Name],0)),0)'
        else:
            prior = f"{week_col(w-1)}{rr}"
        # T_Shipments(Grid_Incoming)は、TTAF供給材料については「TTAFが外部の仕入先から新しく
        # 仕入れてTTAF倉庫に到着する」実績・予定を表す。これはTTAF倉庫内で場所を移しただけの
        # 動きではなく、純粋に合計在庫へ新規に入ってくる量なので、TTAF供給材料かどうかで
        # 特別扱いする必要はなく、他の材料と同じ「前週+入庫-消費」で計算してよい。
        normal = f"{prior}+'Grid_Incoming'!{cl}{rr}-'Grid_Requirement'!{cl}{rr}"
        # 優先順位: 手動棚卸(T_StockCount) > 自社+TTAF実績の合計(両方揃っている週のみ) > 通常のロールフォワード
        ws_st.cell(row=rr, column=1 + w).value = (
            f"=IF({has_count}>0,{count_val},"
            f"IF(({has_self})*({has_ttaf})>0,{self_val}+{ttaf_val},{normal}))"
        )

        # Grid_TheoreticalStock: T_StockCount・自社/TTAF実績を一切見ない、純粋な
        # 「前週+入庫-消費」のロールフォワードのみ(Grid_Stockの優先順位チェーンとは無関係に、
        # 常にこの計算式だけを使う)。Dashboardの「理論在庫」行として、実際の値(Grid_Stock=
        # 「実在庫」行)との乖離を確認するために参照する。
        if w == 1:
            theo_prior = f'IFERROR(INDEX(T_OpeningStock[Opening_Qty],MATCH($A{rr},T_OpeningStock[Part Name],0)),0)'
        else:
            theo_prior = f"{week_col(w-1)}{rr}"
        ws_theo.cell(row=rr, column=1 + w).value = (
            f"={theo_prior}+'Grid_Incoming'!{cl}{rr}-'Grid_Requirement'!{cl}{rr}"
        )

n = len(rm_master) + 1
for ws_ in (ws_req, ws_in, ws_st, ws_theo):
    style_header(ws_, N_WEEKS + 1)
    add_table(ws_, ws_.title, f"A1:{week_col(N_WEEKS)}{n}", style="TableStyleMedium2")
    ws_.column_dimensions["A"].width = 14
    for w in range(1, N_WEEKS + 1):
        ws_.column_dimensions[week_col(w)].width = 9
    ws_.freeze_panes = "B2"

# ============================================================ Material_Detail
# 「どの材料が何に使われているか」を、材料ごとにブロックで見せるトレーサビリティ表示。
# 90枚のシートに分けず、1枚のシートに材料ブロックを縦に並べる（重さの原因を再現しない）。
bom_by_rm = {}
for r in bom:
    bom_by_rm.setdefault(r["RM_Code"], []).append(r)

ws = wb.create_sheet("Material_Detail")
# 選択週の値をピン留め列に複製する方式は廃止。ピン留め列があった位置(旧D列)は削除し、
# 週データの本物の列(D列〜)をそのままA〜C列の直後に配置。選択週(C1)を入力すると、
# VBAのWorksheet_Change(macros/RefreshData_Display.basのJumpToSelectedWeekを呼び出す。導入は
# 任意・手動設定)がウィンドウを横スクロールし、本物の該当週列を固定ペインのすぐ右に
# 表示する（複製データではないため、Dashboard等との数値の食い違いが原理的に起こらない）。
WEEK_START_COL = 4  # column D
def mdetail_week_col(w):
    return get_column_letter(WEEK_START_COL + w - 1)
# 週データ列の1つ右: 「No. of batches」行がPP_Grid内の何行目に対応するかを1回だけMATCHして
# キャッシュしておく内部ヘルパー列(M_BOMのPPGridRowと同じ考え方)。週ごとに毎回MATCHし直すと
# 711(BOMペア数)×104週分のMATCHが発生し重くなるため、行位置は1回だけ求めてINDEXで使い回す
# （列位置はPP_Gridの列並びが固定のため、週番号からそのままw+1として直接指定できる）。
HELPER_COL_MD = WEEK_START_COL + N_WEEKS
# HELPER_COL_MDのもう1つ右: 「Order(発注予定,kg)」行にだけPart Nameを複製しておく見えない
# 列。PO_Draft_*シートがMATCHでこの行を直接特定し、手入力された発注数量を拾うために使う
# (ブロックの長さが材料ごとに違うため、ヘッダー行から固定オフセットではOrder行を特定できない)。
MD_ORDER_HELPER_COL = HELPER_COL_MD + 1
md_order_helper_col_letter = get_column_letter(MD_ORDER_HELPER_COL)
md_week_first_col_letter = get_column_letter(WEEK_START_COL)
md_week_last_col_letter = get_column_letter(WEEK_START_COL + N_WEEKS - 1)

MD_MONTHYEAR_ROW, MD_DATE_ROW, MD_WEEKNO_ROW, MD_TABLE_ROW = 3, 4, 5, 6

ws["A1"] = "選択週を入力（例: W23。現在年の週Noで検索します）"
ws["A1"].font = Font(bold=True)
ws["C1"] = ""
ws["C1"].fill = INPUT_FILL
ws["C1"].font = Font(bold=True, size=12)

_md_cal_first = CAL_HEADER_ROW + 1
_md_cal_last = CAL_HEADER_ROW + N_WEEKS
_md_cal_year_rng = f"Cal_Weeks!$C${_md_cal_first}:$C${_md_cal_last}"
_md_cal_weekofyear_rng = f"Cal_Weeks!$D${_md_cal_first}:$D${_md_cal_last}"
_md_cal_weekindex_rng = f"Cal_Weeks!$A${_md_cal_first}:$A${_md_cal_last}"
_md_cal_label_rng = f"Cal_Weeks!$E${_md_cal_first}:$E${_md_cal_last}"
_md_cal_weekstart_rng = f"Cal_Weeks!$B${_md_cal_first}:$B${_md_cal_last}"
_md_cal_monthyear_rng = f"Cal_Weeks!$G${_md_cal_first}:$G${_md_cal_last}"

# Dashboardと同じ仕組み(TODAY()等の揮発性関数を使わず、AnchorYearを「現在年」として使う
# SUMPRODUCT)。DashboardのC1とは独立した、このシート専用の入力。
_md_wk_match = (f"({_md_cal_year_rng}=Cal_Weeks!$B$1)*"
                f'({_md_cal_weekofyear_rng}=VALUE(SUBSTITUTE(UPPER(TRIM($C$1)),"W","")))')
ws["F1"] = (
    f'=IFERROR(IF(SUMPRODUCT({_md_wk_match})=0,"",SUMPRODUCT({_md_wk_match}*({_md_cal_weekindex_rng}))),"")'
)
ws["F1"].font = Font(size=8, color="808080")
ws["E1"] = "→WeekIndex"
ws["E1"].font = Font(size=8, color="808080")
ws["D1"] = (
    f'=IF($C$1="","週Noを入力してください（例: W23）",'
    f'IF($F$1="","該当週が見つかりません（今年の週Noか確認してください）",'
    f'INDEX({_md_cal_label_rng},$F$1)&" が該当週です（VBA導入時は自動でスクロールします）"))'
)
ws["D1"].font = Font(bold=True, color="0563C1")

for w in range(1, N_WEEKS + 1):
    col = WEEK_START_COL + w - 1
    cal_row = CAL_HEADER_ROW + w
    if w == 1:
        my_formula = f"='Cal_Weeks'!G{cal_row}"
    else:
        prev_cal_row = CAL_HEADER_ROW + w - 1
        my_formula = (
            f'=IF(TEXT(\'Cal_Weeks\'!B{cal_row},"mmm-yy")<>TEXT(\'Cal_Weeks\'!B{prev_cal_row},"mmm-yy"),'
            f"'Cal_Weeks'!G{cal_row},\"\")"
        )
    ws.cell(row=MD_MONTHYEAR_ROW, column=col, value=my_formula)
    dcell = ws.cell(row=MD_DATE_ROW, column=col, value=f"='Cal_Weeks'!B{cal_row}")
    dcell.number_format = "m/d"
    ws.cell(row=MD_WEEKNO_ROW, column=col, value=f"='Cal_Weeks'!D{cal_row}")
    ws.cell(row=MD_TABLE_ROW, column=col, value=f"='Cal_Weeks'!E{cal_row}")
last_col_md = WEEK_START_COL + N_WEEKS - 1
for r in (MD_MONTHYEAR_ROW, MD_DATE_ROW, MD_WEEKNO_ROW, MD_TABLE_ROW):
    for c in range(1, last_col_md + 1):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="D9E1F2")
        ws.cell(row=r, column=c).font = Font(bold=(r in (MD_MONTHYEAR_ROW, MD_TABLE_ROW)))
ws.cell(row=MD_TABLE_ROW, column=1, value="Part Name")
ws.cell(row=MD_TABLE_ROW, column=2, value="項目")
ws.cell(row=MD_TABLE_ROW, column=3, value="1バッチ使用量(kg)")

row_num = MD_TABLE_ROW
# ブロックの並び順は、bom.csv内での初出順ではなく、rm_master(Dashboard等と同じ並び順、
# Substrate→その他Chemical→Ester Film/PP Film→TPZ系)に揃える。BOM行が1つも無い材料
# (Original Towel等)はこれまで通りブロックを作らない。
for _r in rm_master:
    rm_code = _r["RM_Code"]
    entries = bom_by_rm.get(rm_code)
    if not entries:
        continue
    grow = rm_row[rm_code]
    desc = _r["Description"]

    row_num += 1
    mat_header_row = row_num
    ws.cell(row=row_num, column=1, value=rm_code)
    ws.cell(row=row_num, column=2, value=desc)
    for c in range(1, last_col_md + 1):
        ws.cell(row=row_num, column=c).fill = PatternFill("solid", fgColor="FFE699")
        ws.cell(row=row_num, column=c).font = Font(bold=True)
    # MOQ(最小発注量)は数式化せず、手入力用のセルとして空けておく
    moq_cell = ws.cell(row=row_num, column=3, value=None)
    moq_cell.fill = INPUT_FILL
    moq_cell.font = Font(bold=False)
    moq_cell.comment = Comment("MOQ(最小発注量)を入力してください（手書きでOK）", "SOH")

    for entry in entries:
        inter = entry["Intermediate"]

        row_num += 1
        batches_row = row_num
        ws.cell(row=row_num, column=2, value=inter)
        ws.cell(row=row_num, column=3, value="No. of batches")
        helper_col_letter = get_column_letter(HELPER_COL_MD)
        ws.cell(row=row_num, column=HELPER_COL_MD,
                value=f'=IFERROR(MATCH($B{row_num},PP_Grid[Intermediate],0),99999)')
        ws.cell(row=row_num, column=HELPER_COL_MD).font = Font(size=8, color="808080")
        for w in range(1, N_WEEKS + 1):
            cell = ws.cell(row=row_num, column=WEEK_START_COL + w - 1)
            cell.value = (
                f'=IFERROR(INDEX(PP_Grid[#Data],${helper_col_letter}{row_num},{w + 1}),0)'
            )

        row_num += 1
        ws.cell(row=row_num, column=2, value="使用量(kg)")
        ws.cell(row=row_num, column=3,
                value=f'=SUMIFS(M_BOM[RM_Qty_Per_Batch],M_BOM[Intermediate],$B{row_num-1},M_BOM[Part Name],$A{mat_header_row})')
        for w in range(1, N_WEEKS + 1):
            wc = mdetail_week_col(w)
            ws.cell(row=row_num, column=WEEK_START_COL + w - 1,
                    value=f"=$C{row_num}*{wc}{batches_row}")

    row_num += 1
    ws.cell(row=row_num, column=2, value="合計使用量(kg)/週")
    ws.cell(row=row_num, column=2).font = Font(bold=True)
    for w in range(1, N_WEEKS + 1):
        wc = mdetail_week_col(w)
        ws.cell(row=row_num, column=WEEK_START_COL + w - 1, value=f"='Grid_Requirement'!{week_col(w)}{grow}")

    # TTAF在庫(実績)/自社在庫(実績)/合計在庫(週末時点) の3行を追加。
    # 合計在庫は「前週在庫-使用量+TTAF+自社+入庫」を再度ここで組み立てるのではなく、
    # 既にDashboardで検証済みの優先順位ロジック(手動棚卸 > 自社+TTAF実績 > ロールフォワード)
    # を持つGrid_Stockをそのまま参照する(Dashboardと数字が食い違うのを防ぐため)。
    ss_row = ss_row_map[rm_code]
    row_num += 1
    ws.cell(row=row_num, column=2, value="TTAF在庫(実績,kg)")
    for w in range(1, N_WEEKS + 1):
        ws.cell(row=row_num, column=WEEK_START_COL + w - 1,
                value=f"='T_TTAFStock'!{week_col(w)}{ss_row}")

    row_num += 1
    ws.cell(row=row_num, column=2, value="自社在庫(実績,kg)")
    for w in range(1, N_WEEKS + 1):
        ws.cell(row=row_num, column=WEEK_START_COL + w - 1,
                value=f"='T_SelfStock'!{week_col(w)}{ss_row}")

    # Order行: 発注予定数量を材料×週で直接手入力する欄(黄色の入力セル)。以前はT_PlannedOrders
    # を参照する数式だったが、PO_Draft側の自動発注計算をやめ手入力に統一したのに合わせて、
    # ここも生の数値を直接入力する形に変更した。PO_Draft_*シートはMD_ORDER_HELPER_COL列
    # (Part Nameをこの行にだけ複製した見えない列)を目印に、ここで入力した値をそのまま
    # INDEX/MATCHで拾って発注書のレイアウトに転記する(発注書側では計算をしない)。
    row_num += 1
    order_row = row_num
    ws.cell(row=row_num, column=2, value="Order(発注予定,kg)")
    ws.cell(row=row_num, column=MD_ORDER_HELPER_COL, value=rm_code)
    ws.cell(row=row_num, column=MD_ORDER_HELPER_COL).font = Font(size=8, color="808080")
    for w in range(1, N_WEEKS + 1):
        cell = ws.cell(row=row_num, column=WEEK_START_COL + w - 1, value=None)
        cell.fill = INPUT_FILL

    row_num += 1
    ws.cell(row=row_num, column=2, value="合計在庫(週末時点,kg)")
    ws.cell(row=row_num, column=2).font = Font(bold=True)
    for w in range(1, N_WEEKS + 1):
        ws.cell(row=row_num, column=WEEK_START_COL + w - 1,
                value=f"='Grid_Stock'!{week_col(w)}{grow}")

    row_num += 1
    ws.cell(row=row_num, column=2, value="（発注の目安はDashboardの基準在庫[下限/上限]と色分けを参照）")
    ws.cell(row=row_num, column=2).font = Font(italic=True, color="808080")

    row_num += 1  # blank separator row

last_row = row_num
ws.column_dimensions["A"].width = 12
ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 14
for w in range(1, N_WEEKS + 1):
    ws.column_dimensions[mdetail_week_col(w)].width = 9
ws.column_dimensions[get_column_letter(HELPER_COL_MD)].width = 10
ws.column_dimensions[get_column_letter(MD_ORDER_HELPER_COL)].width = 14
ws.freeze_panes = f"{get_column_letter(WEEK_START_COL)}{MD_TABLE_ROW+1}"

# 注: Dashboardにある選択週の列を条件付き書式で強調する仕組みはここでは追加していません。
# Material_Detailは行数が約1,660行と多く、週列全体(104週)に条件付き書式を適用すると
# 対象セル数がDashboardの約16倍(17万セル超)になり、まさに今回のパフォーマンス問題と
# 同種のリスクを新たに持ち込むことになるため。C1入力→VBA(JumpToSelectedWeek)による
# 自動横スクロールのみで十分実用的です。
print("Material_Detail: blocks for", len(bom_by_rm), "materials,", last_row, "rows")

# ============================================================ Dashboard
# 「最終的にここで在庫を確認する」メイン画面。原材料×週の在庫を2年分横軸で見渡せる。
# A〜H列(RM情報。基準在庫の下限・上限を含む)を固定し、その右に本物の週データ列(I列〜)を
# そのまま並べる（選択週の値を複製するピン留め列は廃止）。C1に週No(例:W23)を入力すると、
# VBAのWorksheet_Change(macros/RefreshData_Display.basのJumpToSelectedWeekを呼び出す。導入は
# 任意・手動設定)がウィンドウを横スクロールし、本物の該当週列を固定ペインのすぐ右に
# 表示する。複製データではないため、条件付き書式や数値がずれる余地がない。
# Statusのテキスト列は廃止し、各週のセル自体を基準在庫の下限/上限に対して
# 赤(下限未満)/緑(範囲内)/青(上限超)に色分けする方式にした。
#
# T_SelfStock/T_TTAFStockが「材料×週」のグリッド形式になったため、直近実績の検索は
# 「その材料の行(週1〜週104)の中で一番右にある空欄でないセル」をLOOKUPの最終一致トリックで
# 探すだけで済む(以前のような、行数が育つ長い列を毎回$12000行スキャンする必要がなくなった)。
#
# 【理論在庫／実在庫の2段表示について】材料ごとに2行(上段=理論在庫、下段=実在庫)を並べる。
# 「実在庫」行は従来どおりGrid_Stock(手動棚卸 > 自社+TTAF実績 > ロールフォワード、の優先順位)を
# 参照する。「理論在庫」行は新設のGrid_TheoreticalStock(棚卸・実績を一切見ない、常に純粋な
# ロールフォワードのみ)を参照する。両者の差(乖離)が大きいほど、システム上の計算(原単位・
# 生産計画等)と実際の現場がズレていることを意味する。基準在庫の赤/緑/青の色分けは「実在庫」
# 行のみに適用する(発注判断に使うのは実際の在庫のため。理論在庫行は参考表示)。
LEFT_COLS = ["Part Name", "Description", "Category", "基準在庫_下限", "基準在庫_上限",
             "自社在庫(実績)", "TTAF在庫(実績)", "実績週", "行種別", "乖離(kg)"]
WEEK_START_COL_DASH = len(LEFT_COLS) + 1  # K列から週データ
DASH_ROW_LABEL_COL = 9   # 行種別(理論在庫/実在庫)
DASH_DIFF_COL = 10       # 乖離(kg)
HDR_MONTHYEAR_ROW = 3
HDR_DATE_ROW = 4
HDR_WEEKNO_ROW = 5
HDR_TABLE_ROW = 6  # Excel Table の見出し行としても使う(Label)
DATA_START_ROW = 7

ws = wb.create_sheet("Dashboard")
ws["A1"] = "選択週を入力（例: W23。現在年の週Noで検索します）"
ws["A1"].font = Font(bold=True)
ws["C1"] = ""
ws["C1"].fill = INPUT_FILL
ws["C1"].font = Font(bold=True, size=12)

cal_data_first = CAL_HEADER_ROW + 1
cal_data_last = CAL_HEADER_ROW + N_WEEKS
cal_year_rng = f"Cal_Weeks!$C${cal_data_first}:$C${cal_data_last}"
cal_weekofyear_rng = f"Cal_Weeks!$D${cal_data_first}:$D${cal_data_last}"
cal_weekindex_rng = f"Cal_Weeks!$A${cal_data_first}:$A${cal_data_last}"
cal_label_rng = f"Cal_Weeks!$E${cal_data_first}:$E${cal_data_last}"
cal_weekstart_rng = f"Cal_Weeks!$B${cal_data_first}:$B${cal_data_last}"
cal_monthyear_rng = f"Cal_Weeks!$G${cal_data_first}:$G${cal_data_last}"

# 入力("W23"/"w23"/"23"等)から週Noを取り出し、「現在年(Cal_WeeksのB1=AnchorYear)」×週Noに
# 一致するCal_WeeksのWeekIndexをSUMPRODUCTで求める。MAXIFS/LOOKUPとテーブル構造化参照の組み合わせは
# 本環境で不具合を確認済みのため、ここではプレーン範囲のSUMPRODUCTを使用（マクロ不要）。
# TODAY()等の揮発性関数はワークブック全体の再計算負荷を増やすため使わず、既存のAnchorYear
# セル（週No.のリセット基準としてすでに使われている「現在の基準年」）をそのまま流用する。
_wk_match = (f"({cal_year_rng}=Cal_Weeks!$B$1)*"
             f'({cal_weekofyear_rng}=VALUE(SUBSTITUTE(UPPER(TRIM($C$1)),"W","")))')
ws["F1"] = (
    # 該当週が0件(存在しない週No等)の場合はSUMPRODUCTが0を返しIFERRORでは捕捉できないため、
    # 一致件数を先に判定してから空文字を返す。
    f'=IFERROR(IF(SUMPRODUCT({_wk_match})=0,"",SUMPRODUCT({_wk_match}*({cal_weekindex_rng}))),"")'
)
ws["F1"].font = Font(size=8, color="808080")
ws["E1"] = "→WeekIndex"
ws["E1"].font = Font(size=8, color="808080")
ws["D1"] = (
    f'=IF($C$1="","週Noを入力してください（例: W23）",'
    f'IF($F$1="","該当週が見つかりません（今年の週Noか確認してください）",'
    f'INDEX({cal_label_rng},$F$1)&" が該当週です（VBA導入時は自動でスクロールします）"))'
)
ws["D1"].font = Font(bold=True, color="0563C1")

for w in range(1, N_WEEKS + 1):
    col = WEEK_START_COL_DASH + w - 1
    cal_row = CAL_HEADER_ROW + w
    cl = get_column_letter(col)
    if w == 1:
        my_formula = f"='Cal_Weeks'!G{cal_row}"
    else:
        prev_cal_row = CAL_HEADER_ROW + w - 1
        my_formula = (
            f'=IF(TEXT(\'Cal_Weeks\'!B{cal_row},"mmm-yy")<>TEXT(\'Cal_Weeks\'!B{prev_cal_row},"mmm-yy"),'
            f"'Cal_Weeks'!G{cal_row},\"\")"
        )
    ws.cell(row=HDR_MONTHYEAR_ROW, column=col, value=my_formula)
    dcell = ws.cell(row=HDR_DATE_ROW, column=col, value=f"='Cal_Weeks'!B{cal_row}")
    dcell.number_format = "m/d"
    ws.cell(row=HDR_WEEKNO_ROW, column=col, value=f"='Cal_Weeks'!D{cal_row}")
    ws.cell(row=HDR_TABLE_ROW, column=col, value=f"='Cal_Weeks'!E{cal_row}")
    ws.column_dimensions[cl].width = 9

for c, name in enumerate(LEFT_COLS, start=1):
    ws.cell(row=HDR_TABLE_ROW, column=c, value=name)

style_header(ws, WEEK_START_COL_DASH + N_WEEKS - 1, row=HDR_TABLE_ROW)
for row_i in (HDR_MONTHYEAR_ROW, HDR_DATE_ROW, HDR_WEEKNO_ROW):
    for c in range(1, WEEK_START_COL_DASH + N_WEEKS):
        ws.cell(row=row_i, column=c).fill = PatternFill("solid", fgColor="D9E1F2")
        ws.cell(row=row_i, column=c).font = Font(bold=(row_i == HDR_MONTHYEAR_ROW))

for col, w in zip("ABCDEFGHIJ", [14, 32, 16, 12, 12, 12, 12, 10, 10, 12]):
    ws.column_dimensions[col].width = w

last_col_dash = get_column_letter(WEEK_START_COL_DASH + N_WEEKS - 1)
for i, r in enumerate(rm_master):
    # 材料ごとに2行(理論在庫→実在庫の順)。Part Name(A列)は両方の行に同じ値を書く。
    # RemoveMaterialのDeleteMatchingGridRowはA列の一致だけで行削除するため、これにより
    # 2行とも自動的にまとめて削除される(VBA側の特別対応は不要)。
    theo_rr = DATA_START_ROW + i * 2
    actual_rr = theo_rr + 1
    grow = rm_row[r["RM_Code"]]
    rm = r["RM_Code"]
    ss_row = ss_row_map[r["RM_Code"]]
    ss_first_col = week_col(1)
    ss_last_col = week_col(N_WEEKS)
    ss_self_rng = f"'T_SelfStock'!${ss_first_col}${ss_row}:${ss_last_col}${ss_row}"
    ss_ttaf_rng = f"'T_TTAFStock'!${ss_first_col}${ss_row}:${ss_last_col}${ss_row}"
    ss_label_rng = f"'T_SelfStock'!${ss_first_col}${SS_TABLE_ROW}:${ss_last_col}${SS_TABLE_ROW}"
    # 乖離(kg) = 実在庫(Grid_Stock) - 理論在庫(Grid_TheoreticalStock)。表示期間の最終週の値どうしを
    # 比較するだけでよい。理由: 手動棚卸(T_StockCount)・自社/TTAF実績のどちらで補正されても、
    # 補正が入った週以降は両シートとも「補正後の値+入庫-消費」を同じように積み上げていくため、
    # 補正で生じた差分(乖離)はその週以降ずっと一定のまま変わらない(補正が無ければ差は常に0)。
    # そのため、直近の実績週を特定するLOOKUP等を使わずに、表示している週の範囲でいちばん先
    # (最終週)の値を単純に引き算するだけで、現時点までの累積乖離を取り出せる。
    gs_last_col = week_col(N_WEEKS)
    diff_formula = f"='Grid_Stock'!{gs_last_col}{grow}-'Grid_TheoreticalStock'!{gs_last_col}{grow}"
    for rr in (theo_rr, actual_rr):
        ws.cell(row=rr, column=1, value=rm)
        ws.cell(row=rr, column=2,
                value=f'=IFERROR(INDEX(M_RawMaterials[Description],MATCH($A{rr},M_RawMaterials[Part Name],0)),"")')
        ws.cell(row=rr, column=3,
                value=f'=IFERROR(INDEX(M_RawMaterials[Category],MATCH($A{rr},M_RawMaterials[Part Name],0)),"")')
        ws.cell(row=rr, column=4,
                value=f'=IFERROR(INDEX(M_RawMaterials[基準在庫下限_要入力],MATCH($A{rr},M_RawMaterials[Part Name],0)),0)')
        ws.cell(row=rr, column=5,
                value=f'=IFERROR(INDEX(M_RawMaterials[基準在庫上限_要入力],MATCH($A{rr},M_RawMaterials[Part Name],0)),0)')
        ws.cell(row=rr, column=6,
                value=f'=IFERROR(LOOKUP(2,1/({ss_self_rng}<>""),{ss_self_rng}),"")')
        ws.cell(row=rr, column=7,
                value=f'=IFERROR(LOOKUP(2,1/({ss_ttaf_rng}<>""),{ss_ttaf_rng}),"")')
        ws.cell(row=rr, column=8,
                value=f'=IFERROR(LOOKUP(2,1/({ss_self_rng}<>""),{ss_label_rng}),"")')
        ws.cell(row=rr, column=DASH_DIFF_COL, value=diff_formula)
    ws.cell(row=theo_rr, column=DASH_ROW_LABEL_COL, value="理論在庫")
    ws.cell(row=theo_rr, column=DASH_ROW_LABEL_COL).font = Font(italic=True, color="808080")
    ws.cell(row=actual_rr, column=DASH_ROW_LABEL_COL, value="実在庫")
    ws.cell(row=actual_rr, column=DASH_ROW_LABEL_COL).font = Font(bold=True)
    for w in range(1, N_WEEKS + 1):
        col = WEEK_START_COL_DASH + w - 1
        gs_col = week_col(w)
        ws.cell(row=theo_rr, column=col, value=f"='Grid_TheoreticalStock'!{gs_col}{grow}")
        ws.cell(row=theo_rr, column=col).font = Font(italic=True, color="808080")
        ws.cell(row=actual_rr, column=col, value=f"='Grid_Stock'!{gs_col}{grow}")

n_last_row = DATA_START_ROW + len(rm_master) * 2 - 1
# ヘッダー行が数式(Cal_Weeks参照)のため、Excelのテーブル機能(ListObject)ではなく
# 罫線・縞模様の手動書式で「テーブルらしい」見た目にする（テーブル見出しは文字列必須のため）。
# 【補足】材料ごとに理論在庫→実在庫の2行が並ぶ構成のため、この縞模様(1行おき)は偶然にも
# 「理論在庫行(奇数オフセット)は常にグレー、実在庫行(偶数オフセット)は常に白」という、
# 2種類の行を視覚的に区別する効果も兼ねている。
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for rr2 in range(HDR_TABLE_ROW, n_last_row + 1):
    stripe = (rr2 - HDR_TABLE_ROW) % 2 == 1
    for c in range(1, WEEK_START_COL_DASH + N_WEEKS):
        cell = ws.cell(row=rr2, column=c)
        cell.border = border
        if rr2 > HDR_TABLE_ROW and stripe:
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
ws.freeze_panes = f"{get_column_letter(WEEK_START_COL_DASH)}{DATA_START_ROW}"

# 基準在庫の下限・上限に対して、各週のセルを赤(下限未満)/緑(範囲内)/青(上限超)に色分け
# （$D=基準在庫_下限、$E=基準在庫_上限。行内の相対参照なのでどの週列でも自分の行の基準を見る）。
# $I(行種別)="実在庫"の行だけに適用する(発注判断に使うのは実際の在庫のため。理論在庫行は
# グレー文字の参考表示のみで、色分けの対象にはしない)。
row_label_col_letter = get_column_letter(DASH_ROW_LABEL_COL)
stock_band_range = f"{get_column_letter(WEEK_START_COL_DASH)}{DATA_START_ROW}:{last_col_dash}{n_last_row}"
ws.conditional_formatting.add(
    stock_band_range,
    FormulaRule(formula=[f'AND(${row_label_col_letter}{DATA_START_ROW}="実在庫",'
                          f"{get_column_letter(WEEK_START_COL_DASH)}{DATA_START_ROW}<$D{DATA_START_ROW})"],
                fill=PatternFill("solid", fgColor="FFC7CE"))  # 赤: 基準在庫の下限未満
)
ws.conditional_formatting.add(
    stock_band_range,
    FormulaRule(formula=[f'AND(${row_label_col_letter}{DATA_START_ROW}="実在庫",'
                          f"{get_column_letter(WEEK_START_COL_DASH)}{DATA_START_ROW}>=$D{DATA_START_ROW},"
                          f"{get_column_letter(WEEK_START_COL_DASH)}{DATA_START_ROW}<=$E{DATA_START_ROW})"],
                fill=PatternFill("solid", fgColor="C6EFCE"))  # 緑: 基準在庫の範囲内
)
ws.conditional_formatting.add(
    stock_band_range,
    FormulaRule(formula=[f'AND(${row_label_col_letter}{DATA_START_ROW}="実在庫",'
                          f"{get_column_letter(WEEK_START_COL_DASH)}{DATA_START_ROW}>$E{DATA_START_ROW})"],
                fill=PatternFill("solid", fgColor="BDD7EE"))  # 青: 基準在庫の上限超
)
# 選択中の週(F1のWeekIndex)に該当する列を太枠で強調（COLUMN()の自己参照なので
# テーブル構造化参照やLOOKUP配列を使わずに済み、本環境で確認済みの不具合を回避できる）。
# 上の赤/緑/青の塗りつぶしと競合しないよう、塗りつぶしではなく罫線でハイライトする。
week_select_border = Border(left=Side(style="thick", color="BF8F00"), right=Side(style="thick", color="BF8F00"))
ws.conditional_formatting.add(
    f"{get_column_letter(WEEK_START_COL_DASH)}{HDR_TABLE_ROW}:{last_col_dash}{n_last_row}",
    FormulaRule(
        formula=[f'(COLUMN()-COLUMN(${get_column_letter(WEEK_START_COL_DASH)}$1)+1)=$F$1'],
        border=week_select_border,
    )
)
print("Dashboard: week-by-week grid for", len(rm_master), "materials x", N_WEEKS, "weeks")

# ============================================================ PO Draft sheets (Chemical Release format)
# 注文書は2年先まで不要。翌月(Firm)＋翌々月・翌々翌月(Forecast)の3ヶ月＝約13週分だけを表示する。
# 表示開始週はP7セル(基準週WeekIndex)で指定し、ここを変えるだけで対象月をずらせる（数式のみ、マクロ不要）。
PO_N_WEEKS = 13
PO_HDR_MONTHYEAR_ROW = 10
PO_HDR_DATE_ROW = 11
PO_HDR_WEEKNO_ROW = 12
PO_HDR_FIRMFORECAST_ROW = 13
PO_HDR_TABLE_ROW = 14
PO_DATA_START_ROW = 15
PO_FIRST_WEEK_COL = 8  # H列


def build_po_draft(sheet_name, category, title):
    ws = wb.create_sheet(sheet_name)
    ws["B2"] = "TO：（サプライヤー／TTAF担当者名を入力）"
    ws["B3"] = "　　　　　（会社名）"
    ws["B5"] = "FROM：（発行者名）"
    ws["B6"] = "　　　　　（自社名）"
    ws["N2"] = "Order Date:"
    ws["P2"] = datetime.date.today()
    ws["N3"] = "Issue Month:"
    ws["N4"] = "Firm Month:"
    ws["N5"] = "Revision"
    ws["P5"] = "00"
    ws["N7"] = "基準週(WeekIndex, 翌月の頭を推奨):"
    # デフォルト値は「来月1日」が属する週のWeekIndex
    default_firm_month = (datetime.date.today().replace(day=1) + datetime.timedelta(days=32)).replace(day=1)
    default_start_week = date_to_week_index(default_firm_month, N_WEEKS)
    ws["P7"] = default_start_week
    ws["P7"].fill = INPUT_FILL
    ws["P7"].font = Font(bold=True)
    ws["B8"] = title
    ws["B8"].font = Font(bold=True, size=12)

    for w in range(1, PO_N_WEEKS + 1):
        col = PO_FIRST_WEEK_COL + w - 1
        if w == 1:
            my_formula = "=INDEX(Cal_Weeks[MonthYearLabel],$P$7)"
        else:
            my_formula = (
                f'=IF(INDEX(Cal_Weeks[MonthYearLabel],$P$7+{w-1})<>INDEX(Cal_Weeks[MonthYearLabel],$P$7+{w-2}),'
                f'INDEX(Cal_Weeks[MonthYearLabel],$P$7+{w-1}),"")'
            )
        ws.cell(row=PO_HDR_MONTHYEAR_ROW, column=col, value=my_formula)
        dcell = ws.cell(row=PO_HDR_DATE_ROW, column=col, value=f"=INDEX(Cal_Weeks[WeekStart],$P$7+{w-1})")
        dcell.number_format = "m/d"
        ws.cell(row=PO_HDR_WEEKNO_ROW, column=col, value=f"=INDEX(Cal_Weeks[WeekOfYear],$P$7+{w-1})")
        ws.cell(row=PO_HDR_FIRMFORECAST_ROW, column=col,
                value=(f'=IF(INDEX(Cal_Weeks[MonthYearLabel],$P$7+{w-1})=INDEX(Cal_Weeks[MonthYearLabel],$P$7),'
                       f'"Firm","Forecast")'))
        ws.cell(row=PO_HDR_TABLE_ROW, column=col, value=f"=INDEX(Cal_Weeks[Label],$P$7+{w-1})")
        ws.column_dimensions[get_column_letter(col)].width = 8

    left_headers = ["Part Name", "TTAF Code", "CSA Code", "UOM", "SafetyStock", "CurrentStock"]
    for c, h in enumerate(left_headers, start=2):
        ws.cell(row=PO_HDR_TABLE_ROW, column=c, value=h)
    total_col = PO_FIRST_WEEK_COL + PO_N_WEEKS
    ws.cell(row=PO_HDR_TABLE_ROW, column=total_col, value="Total")

    last_col_idx = total_col
    for row_i in (PO_HDR_MONTHYEAR_ROW, PO_HDR_DATE_ROW, PO_HDR_WEEKNO_ROW, PO_HDR_FIRMFORECAST_ROW, PO_HDR_TABLE_ROW):
        for c in range(2, last_col_idx + 1):
            cell = ws.cell(row=row_i, column=c)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.font = Font(bold=(row_i in (PO_HDR_MONTHYEAR_ROW, PO_HDR_TABLE_ROW)))

    data_row = PO_HDR_TABLE_ROW
    items = [r for r in rm_master if r["Category"] == category]
    for r in items:
        data_row += 1
        rm = r["RM_Code"]
        # Grid_Stock内の行位置は、固定の数値(grow_rel)ではなくMATCHで毎回動的に求める。
        # 材料の追加・削除でGrid_Stockの行位置がずれても(AddMaterial/RemoveMaterialマクロ参照)、
        # 数式側が自動的に正しい行を追従できるようにするため。
        grow_match = f'MATCH($D{data_row},Grid_Stock[Part Name],0)'
        ws.cell(row=data_row, column=2, value=f'=IFERROR(INDEX(M_RawMaterials[Description],MATCH("{rm}",M_RawMaterials[Part Name],0)),"")')
        ws.cell(row=data_row, column=3, value=r.get("TTAF_Code", ""))
        ws.cell(row=data_row, column=4, value=rm)
        ws.cell(row=data_row, column=5, value="kg")
        ws.cell(row=data_row, column=6, value=f'=IFERROR(INDEX(M_RawMaterials[基準在庫下限_要入力],MATCH("{rm}",M_RawMaterials[Part Name],0)),0)')
        ws.cell(row=data_row, column=7, value=f"=INDEX(Grid_Stock[#Data],{grow_match},$P$7)")
        # 発注数量は自動計算せず、Material_DetailのOrder(発注予定,kg)行に手入力された値を
        # そのまま転記するだけ。MD_ORDER_HELPER_COL列(Order行にだけPart Nameが複製されている
        # 見えない列)をMATCHでたどることで、材料ごとにブロックの長さが違っても正しい行を
        # 一意に特定できる。
        md_order_match = f'MATCH($D{data_row},Material_Detail!${md_order_helper_col_letter}:${md_order_helper_col_letter},0)'
        for w in range(1, PO_N_WEEKS + 1):
            col = PO_FIRST_WEEK_COL + w - 1
            ws.cell(row=data_row, column=col,
                    value=f"=IFERROR(INDEX(Material_Detail!${md_week_first_col_letter}:${md_week_last_col_letter},{md_order_match},$P$7+{w-1}),0)")
        rng_start = get_column_letter(PO_FIRST_WEEK_COL)
        rng_end = get_column_letter(PO_FIRST_WEEK_COL + PO_N_WEEKS - 1)
        ws.cell(row=data_row, column=total_col, value=f"=SUM({rng_start}{data_row}:{rng_end}{data_row})")

    last_row = data_row
    # ヘッダー行が数式(Cal_Weeks参照)のため、Excelのテーブル機能(ListObject)は使わず
    # 罫線・縞模様の手動書式にする（テーブル見出しは文字列必須で、数式は入れられないため）。
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    if items:
        for rr2 in range(PO_HDR_TABLE_ROW, last_row + 1):
            stripe = (rr2 - PO_HDR_TABLE_ROW) % 2 == 1
            for c in range(2, last_col_idx + 1):
                cell = ws.cell(row=rr2, column=c)
                cell.border = border
                if rr2 > PO_HDR_TABLE_ROW and stripe:
                    cell.fill = PatternFill("solid", fgColor="F2F2F2")
    else:
        ws.cell(row=PO_HDR_TABLE_ROW + 1, column=2, value="(該当品目なし)")
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.freeze_panes = f"{get_column_letter(PO_FIRST_WEEK_COL)}{PO_DATA_START_ROW}"
    print(f"{sheet_name}: {len(items)} items")

build_po_draft("PO_Draft_Chemical", "Chemical", "Chemicals : TTAF Supply")
build_po_draft("PO_Draft_Hazardous", "Hazardous Chemical", "Hazardous Chemicals")
build_po_draft("PO_Draft_Substrate", "Substrate", "Substrates")

# ============================================================ README
ws = wb.create_sheet("README", 0)
readme_lines = [
    "SOH（在庫管理）シート 使い方",
    "",
    "このブックはPower BI等に読み込ませる中間テーブルではなく、これ自体が完成品です。",
    "Dashboard と PO_Draft_* が最終的な閲覧・発注画面、それ以外は計算の土台です。",
    "",
    "【週番号のルール】すべてExcelの数式で計算（Cal_WeeksのB1=AnchorYearを変えると全体が再計算）",
    "  Week1=1月1日を含む月〜日の週。年が変わると再び1にリセットされます（2026-W52の次は2027-W01）。",
    "  各シートの週見出しは3段: 1段目=月-年(Aug-26等) / 2段目=その週月曜の日付 / 3段目=週No。",
    "",
    "【普段見るシート】",
    "  Dashboard           : 原材料×週の在庫を2年分、横軸で見渡せるメイン画面（まずここ）。",
    "                        自社在庫(実績)・TTAF在庫(実績)・実績週の列で内訳も確認できます。",
    "                        基準在庫(下限/上限)を入力すると、各週のセルが赤(下限未満)/",
    "                        緑(範囲内)/青(上限超)に自動で色分けされます。",
    "                        C1に'W23'のように入力すると（現在年の週Noとして検索）、",
    "                        該当する週の列が太枠でハイライトされます。VBA(JumpToSelectedWeek、",
    "                        要ボタン設定・詳細はdocs/SOH_System_Guide.md)を導入していれば、",
    "                        実績週のすぐ右にその週の列が来るよう自動でスクロールもされます。",
    "  Material_Detail     : 材料ごとに「どの中間体が・何バッチ・いくら使うか」をブロック表示（トレーサビリティ）。",
    "                        Dashboardと同様にC1に'W23'のように入力すると該当週が求まり、",
    "                        VBA導入時はその週列のすぐ右に自動でスクロールします。",
    "                        材料名の右のC列にMOQ(最小発注量)を手入力できます。",
    "                        合計使用量の下にTTAF在庫実績・自社在庫実績・Order(発注予定)・",
    "                        合計在庫(週末時点)の週次推移も表示されます。",
    "                        HideInactiveIntermediatesマクロ(要ボタン設定)で、指定期間ずっと生産予定の無い",
    "                        中間体の行を折りたためます（在庫関連の行は常に表示されたままです）。",
    "  PO_Draft_*          : Material_Detailの「Order」行に手入力した発注数量を注文書ひな形に転記",
    "                        （自動計算はしません。発注数量はMaterial_Detail側で入力してください）。",
    "  T_Shipments         : 発注・着荷の入力。TTAF供給材料については「TTAF倉庫への到着実績」を",
    "                        表します(TTAFは仕入先であり倉庫でもあるため)。TTAF以外の材料は",
    "                        従来通り弊社への入庫実績です。RefreshShipmentsでCSA Reportの",
    "                        Shipping Scheduleから一括更新できます（手入力も可）。",
    "  T_OpeningStock/T_StockCount : 入力用",
    "  T_SelfStock/T_TTAFStock : 材料×週のグリッドで自社/TTAF在庫実績を表示（出力・閲覧用）。",
    "                        RefreshSelfStock/RefreshTTAFStockで更新されます。手入力はしないでください",
    "                        （生データは非表示のT_SelfStock_Log/T_TTAFStock_Logに安全に保存されています）。",
    "",
    "  M_RawMaterials・M_BOM・PP_Grid・Grid_Stock・その他非表示シートは内部計算用です。通常は開く必要はありません。",
    "",
    "【重要】原単位・バッチ数・自社/TTAF在庫・発注実績はPythonを使わず、Excel(VBA)マクロだけで更新できます。",
    "  macros/フォルダのRefreshData_*.bas(7ファイル)を導入し、RefreshWeeklyBatches / RefreshBOM / RefreshSelfStock /",
    "  RefreshTTAFStock / RefreshShipments を実行してください（対象ファイルを選ぶだけです。",
    "  詳細はdocs/SOH_System_Guide.md、未検証のため要動作確認）。",
    "  T_Shipments・T_OpeningStock・T_StockCount等の入力内容は上書きされません",
    "  （T_ShipmentsはRefreshShipmentsを実行した場合のみ更新されます）。",
    "  Material_Detailの中間体の行を隠す/戻すHideInactiveIntermediates / ShowAllIntermediatesは、",
    "  ボタンへの割り当てが必要な一度だけの手動設定です（詳細はdocs/SOH_System_Guide.md）。",
    "",
    "【毎月の運用】",
    "  1. 「Powder & Slurry & Pgm Plan」の新しい月版でRefreshWeeklyBatchesを実行",
    "  2. 「Raw Material - Look Up」が更新されていればRefreshBOMを実行",
    "  3. 自社倉庫の現物確認を毎週月曜の朝に実施したらRefreshSelfStockを実行",
    "  4. CSA Reportが毎週月曜に届いたらRefreshTTAFStockとRefreshShipmentsを実行",
    "  5. 発注する数量はMaterial_Detailの「Order(発注予定,kg)」行に直接入力",
    "  6. 棚卸を実施したらT_StockCountに実測値を追記（Date列に実施日を入力。WeekIndex列は自動計算）",
    "  7. Dashboardで赤色(基準在庫の下限未満)の週を確認し、PO_Draft_*から注文書を出力",
    "  8. 月初は、前月最終週と当月頭のDashboardを見比べて在庫差異を確認（Plan Increase and Decrease /",
    "     Inventory Releasesの報告フォーマットに転記）",
    "",
    "【前提・要確認事項】詳細はdocs/SOH_System_Guide.mdを参照",
    "  - M_RawMaterials の基準在庫(下限/上限)とLeadTime_Weeksは仮値(0)です。実際の水準に置き換えてください",
    "    （Dashboardの週次セルの赤/緑/青の色分けに使われます）。",
    "  - Categoryの割り当て(Chemical/Hazardous Chemical/Substrate)は入手データから機械的に推定した部分があります。要レビュー。",
    "  - 週次バッチ数はPowder & Slurry & Pgm Planの実データ（約36材料シートから抽出）を使用しています。",
]
for i, line in enumerate(readme_lines, start=1):
    ws.cell(row=i, column=1, value=line)
ws["A1"].font = Font(bold=True, size=14)
ws.column_dimensions["A"].width = 100

# ============================================================ 操作パネル
# Alt+F8のマクロ一覧は常にマクロ名のアルファベット順に表示され、モジュールの並びや記述順を
# 変えても順序を変えられない(Excelの仕様)。月次の運用手順どおりの順番でマクロを実行したい、
# という要望に対応するため、シート上に手順順のボタン一覧を用意する。openpyxl(このブックの
# 生成に使っているPythonライブラリ)はクリック可能な図形やそこへのマクロ割り当てを自動生成
# できないため、このシートは「どこに・どの順で・どのマクロを割り当てるか」の下地(枠と番号・
# マクロ名・説明)を用意するところまでで、実際の図形の配置とマクロ登録は貴社のExcelで
# 一度だけ手動で行う(下記【ボタンの割り当て方】参照)。
ws_panel = wb.create_sheet("操作パネル")
ws_panel["A1"] = "操作パネル（ボタンの設置場所）"
ws_panel["A1"].font = Font(bold=True, size=14)
ws_panel["A2"] = ("Alt+F8のマクロ一覧は常にアルファベット順になるため、このシートに月次の運用手順の順で"
                  "ボタンを配置します。ボタンの作成・マクロの割り当ては貴社のExcelで一度だけ手動で行って"
                  "ください（下記の番号・マクロ名の行に重ねて図形を描き、「マクロの登録」で対応するマクロ名を"
                  "選ぶだけです）。詳しい手順はdocs/SOH_System_Guide.mdを参照してください。")
ws_panel["A2"].font = Font(italic=True, color="808080", size=9)
ws_panel["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws_panel.row_dimensions[2].height = 45

PANEL_SLOT_FILL = PatternFill("solid", fgColor="E2EFDA")
PANEL_SECTION_FILL = PatternFill("solid", fgColor="1F4E78")

panel_sections = [
    ("【毎月・毎週の定型作業】上から順に実行", [
        ("RefreshWeeklyBatches", "「Powder & Slurry & Pgm Plan」の新しい月版が出たら実行"),
        ("RefreshBOM", "「Raw Material - Look Up」が更新されたら実行"),
        ("RefreshSelfStock", "自社倉庫の現物確認（daily check）を毎週月曜の朝に実施したら実行"),
        ("RefreshTTAFStock", "CSA Reportが毎週月曜に届いたら実行（TTAF在庫）"),
        ("RefreshShipments", "CSA Reportが毎週月曜に届いたら実行（発注・着荷）"),
    ]),
    ("【任意】Material_Detailの表示調整", [
        ("HideInactiveIntermediates", "しばらく生産予定の無い中間体の行を折りたたむ"),
        ("ShowAllIntermediates", "折りたたんだ行をすべて再表示する"),
    ]),
    ("【まれに使う】材料・中間体の追加/削除（取り消せません。事前にバックアップ推奨）", [
        ("AddMaterial", "新しい材料(TTAF供給品)をシステムに追加する"),
        ("RemoveMaterial", "使わなくなった材料をシステムから削除する"),
        ("RemoveIntermediate", "生産中止になった中間体をシステムから削除する"),
    ]),
]

panel_row = 4
panel_num = 1
for section_title, items in panel_sections:
    ws_panel.cell(row=panel_row, column=1, value=section_title)
    for c in range(1, 4):
        ws_panel.cell(row=panel_row, column=c).fill = PANEL_SECTION_FILL
        ws_panel.cell(row=panel_row, column=c).font = Font(color="FFFFFF", bold=True)
    panel_row += 1
    for macro_name, desc in items:
        ws_panel.cell(row=panel_row, column=1, value=panel_num)
        ws_panel.cell(row=panel_row, column=1).alignment = Alignment(horizontal="center")
        ws_panel.cell(row=panel_row, column=2, value=macro_name)
        ws_panel.cell(row=panel_row, column=2).font = Font(bold=True)
        ws_panel.cell(row=panel_row, column=3, value=desc)
        for c in range(1, 4):
            ws_panel.cell(row=panel_row, column=c).fill = PANEL_SLOT_FILL
            ws_panel.cell(row=panel_row, column=c).border = Border(
                top=Side(style="thin", color="A9A9A9"), bottom=Side(style="thin", color="A9A9A9"),
                left=Side(style="thin", color="A9A9A9"), right=Side(style="thin", color="A9A9A9"))
        panel_row += 1
        panel_num += 1
    panel_row += 1  # セクション間の空白行

ws_panel.column_dimensions["A"].width = 6
ws_panel.column_dimensions["B"].width = 28
ws_panel.column_dimensions["C"].width = 70

# ---- Solution名リスト（RefreshWeeklyBatchesが「Powder & Slurry & Pgm Plan」の行を
# 中間体/完成品(Catalyst)/Solutionに自動判定する際に使う）----
# 中間体(TSP-/TPP-/TSZ-/TVS-/VSP-)・完成品(それ以外)は接頭辞や消去法で機械的に判定できるが、
# Solutionだけは略称(20P・SH等)に共通の接頭辞が無いため、この表で明示的に列挙する。
# 新しいSolutionが増えたら、この表に1行追加するだけでRefreshWeeklyBatchesが自動的に対応する
# （行番号のメンテナンスは一切不要）。
ws_panel["E3"] = "Solution名リスト（RefreshWeeklyBatchesの行判定に使用）"
ws_panel["E3"].font = Font(bold=True)
sol_header_row = 4
ws_panel.cell(row=sol_header_row, column=5, value="略称（Pgm Plan表記）")
ws_panel.cell(row=sol_header_row, column=6, value="正式名（システム内部名）")
solution_aliases = [
    ("20P", "SOL-20P"),
    ("10H", "SOL-10H"),
    ("10H VW", "SOL-10H VW"),
    ("250D", "SOL-250D"),
    ("85D", "SOL-85D"),
    ("SCH", "SOL-SCH"),
    ("SH", "SOL-SH"),
]
for i, (alias, canon) in enumerate(solution_aliases):
    r = sol_header_row + 1 + i
    ws_panel.cell(row=r, column=5, value=alias)
    ws_panel.cell(row=r, column=6, value=canon)
sol_last_row = sol_header_row + len(solution_aliases)
style_header(ws_panel, 2, row=sol_header_row)
for c in (5, 6):
    ws_panel.cell(row=sol_header_row, column=c).fill = HEADER_FILL
    ws_panel.cell(row=sol_header_row, column=c).font = HEADER_FONT
sol_tbl = Table(displayName="T_SolutionNames", ref=f"E{sol_header_row}:F{sol_last_row}")
sol_tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
ws_panel.add_table(sol_tbl)
ws_panel.column_dimensions["E"].width = 22
ws_panel.column_dimensions["F"].width = 22

ws_panel.freeze_panes = "A3"

# ---- ナビゲーション（README上部にジャンプリンクを追加） ----
nav_targets = [
    ("操作パネル", "月次の運用手順順に並んだマクロボタンの設置場所"),
    ("Dashboard", "原材料×週の在庫（2年分・横軸で見渡せるメイン画面。まずここ）"),
    ("Material_Detail", "材料ごとの使用状況（どの材料が何に使われているか）"),
    ("PO_Draft_Chemical", "発注書ドラフト（Chemical）"),
    ("PO_Draft_Hazardous", "発注書ドラフト（Hazardous Chemical）"),
    ("PO_Draft_Substrate", "発注書ドラフト（Substrate）"),
    ("T_Shipments", "発注・着荷の入力（TTAF供給材料はTTAF倉庫への到着実績を表す）"),
    ("T_OpeningStock", "期首在庫の入力"),
    ("T_StockCount", "棚卸実績の入力"),
    ("T_SelfStock", "自社倉庫の在庫実績（材料×週。RefreshSelfStockで自動更新）"),
    ("T_TTAFStock", "TTAF倉庫の在庫実績（材料×週。RefreshTTAFStockで自動更新）"),
]
ws.insert_rows(2, amount=len(nav_targets) + 2)
ws["A2"] = "【ジャンプ】クリックで各シートへ移動"
ws["A2"].font = Font(bold=True, size=12)
for i, (target, label) in enumerate(nav_targets, start=3):
    cell = ws.cell(row=i, column=1, value=f"▶ {target} - {label}")
    cell.hyperlink = f"#'{target}'!A1"
    cell.font = Font(color="0563C1", underline="single")

# ---- 内部処理用シートは非表示にして視認性を上げる（Dashboardが週次在庫の表示を兼ねるためGrid_Stockも非表示。
#      T_SelfStock_Log/T_TTAFStock_LogはVBAが書き込む生ログで、目に見えるT_SelfStock/T_TTAFStock
#      （材料×週のグリッド）はそこから数式で計算するだけなので、生ログ自体は非表示にする） ----
for sheet_name in ["Cal_Weeks", "M_Intermediates", "M_ProductMap", "Grid_Requirement", "Grid_Incoming",
                    "Grid_Stock", "Grid_TheoreticalStock", "T_SelfStock_Log", "T_TTAFStock_Log"]:
    if sheet_name in wb.sheetnames:
        wb[sheet_name].sheet_state = "hidden"

# ---- シートの並び順を業務で使う順に ----
order = ["README", "操作パネル", "Dashboard", "Material_Detail", "PO_Draft_Chemical", "PO_Draft_Hazardous",
         "PO_Draft_Substrate", "T_Shipments", "T_OpeningStock", "T_StockCount",
         "T_SelfStock", "T_TTAFStock",
         "M_RawMaterials", "M_BOM", "PP_Grid",
         "Cal_Weeks", "M_Intermediates", "M_ProductMap", "Grid_Requirement",
         "Grid_Incoming", "Grid_Stock", "Grid_TheoreticalStock", "T_SelfStock_Log", "T_TTAFStock_Log"]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else len(order))

# ---- 保存前チェック: テーブルがヘッダー行のみ(データ0行)になっていないか ----
# (この状態で保存するとExcelで開いたときに「修復」されテーブルが削除されてしまうため)
broken_tables = []
for sheet in wb.worksheets:
    for tbl in sheet.tables.values():
        min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(tbl.ref)
        if max_row <= min_row:
            broken_tables.append(f"{sheet.title}!{tbl.name} (ref={tbl.ref})")
if broken_tables:
    raise RuntimeError(
        "以下のテーブルがヘッダー行のみでデータ行がありません(Excelで開くと修復されます): "
        + ", ".join(broken_tables)
    )

wb.save(OUT_PATH)
print("Full workbook written to", OUT_PATH)
